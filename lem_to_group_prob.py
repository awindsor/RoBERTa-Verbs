#!/usr/bin/env python3
"""
Aggregate group probabilities by lemma (auto-detect group columns).

Input: CSV produced by roberta_mlm_on_verbs.py with group columns appended.

Output:
  - If output filename ends with .csv  → write CSV
  - If output filename ends with .xlsx → write Excel (two sheets)

Excel sheets:
  1) lemma_to_groups:
        lemma → mean group probabilities (formatted as %) + count
        - bold the highest group percentage in each row
        - color the lemma cell BLUE if the 2nd-highest group percentage is at least
          --second-threshold (default 0.50) times the highest
  2) groups_ranked:
        for each group: (lemma, pct) sorted by decreasing pct,
        with the lemma bolded in the group where it has the highest probability

Auto-detection:
  - If --group-cols is provided, use it.
  - Otherwise infer group columns by taking all columns after the last prob_k column.

Requirements for Excel output:
  pip install openpyxl
"""

from __future__ import annotations

import argparse
import csv
import os
import re
from collections import defaultdict
from typing import Dict, List, Tuple

PROB_COL_RE = re.compile(r"^prob_(\d+)$")


# ------------------------- helpers -------------------------

def infer_group_cols(fieldnames: List[str]) -> List[str]:
    """Infer group columns as those appearing after the last prob_k column."""
    last_prob_idx = -1
    for i, name in enumerate(fieldnames):
        if PROB_COL_RE.match(name.strip()):
            last_prob_idx = i
    if last_prob_idx == -1:
        return []
    return [c for c in fieldnames[last_prob_idx + 1:] if c.strip()]


def safe_sheet_title(s: str) -> str:
    bad = r'[]:*?/\\'
    out = "".join("_" if ch in bad else ch for ch in s)
    return out[:31]


# ------------------------- Excel writer -------------------------

def write_excel(
    xlsx_path: str,
    lemma_col: str,
    group_cols: List[str],
    counts: Dict[str, int],
    means: Dict[str, Dict[str, float]],
    best_group_for_lemma: Dict[str, str],
    second_threshold: float,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise SystemExit(
            "Excel output requested but openpyxl is not installed. Install with: pip install openpyxl"
        ) from e

    wb = Workbook()
    bold_font = Font(bold=True)
    blue_font = Font(color="0000FF")  # blue text
    header_font = Font(bold=True)
    header_align = Alignment(vertical="center")

    # ---------- Sheet 1: lemma_to_groups ----------
    ws1 = wb.active
    ws1.title = safe_sheet_title("lemma_to_groups")

    header = [lemma_col] + group_cols + ["count"]
    ws1.append(header)
    for cell in ws1[1]:
        cell.font = header_font
        cell.alignment = header_align

    # Write rows + apply formatting rules
    # Columns:
    #   lemma = 1
    #   groups = 2..(1+len(group_cols))
    #   count = last
    for lemma in sorted(counts):
        row_vals = [lemma] + [means[g][lemma] for g in group_cols] + [counts[lemma]]
        ws1.append(row_vals)
        r = ws1.max_row

        # Determine max/2nd-max across groups for this lemma
        group_vals = [(g, means[g][lemma]) for g in group_cols]
        # stable sort: highest first
        group_vals.sort(key=lambda x: x[1], reverse=True)
        max_g, max_v = group_vals[0]
        second_v = group_vals[1][1] if len(group_vals) > 1 else 0.0

        # Bold the max group cell in this row
        max_idx = group_cols.index(max_g)  # 0-based within group_cols
        max_col = 2 + max_idx  # Excel column index for the group cell
        ws1.cell(row=r, column=max_col).font = bold_font

        # If 2nd-highest >= threshold * highest, color lemma cell blue
        # Avoid dividing by zero: if max_v == 0, treat as "not ambiguous"
        if max_v > 0 and second_v >= (second_threshold * max_v):
            ws1.cell(row=r, column=1).font = blue_font

    # Percent formatting for group columns
    for j in range(2, 2 + len(group_cols)):
        for i in range(2, ws1.max_row + 1):
            ws1.cell(row=i, column=j).number_format = "0.00%"

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(ws1.max_column)}{ws1.max_row}"

    # ---------- Sheet 2: groups_ranked ----------
    ws2 = wb.create_sheet(title=safe_sheet_title("groups_ranked"))

    per_group_sorted: Dict[str, List[Tuple[str, float]]] = {}
    max_len = 0
    for g in group_cols:
        items = [(lemma, means[g][lemma]) for lemma in counts]
        items.sort(key=lambda x: x[1], reverse=True)
        per_group_sorted[g] = items
        max_len = max(max_len, len(items))

    header2: List[str] = []
    for g in group_cols:
        header2.extend([f"{g}_lemma", f"{g}_pct"])
    ws2.append(header2)
    for cell in ws2[1]:
        cell.font = header_font
        cell.alignment = header_align

    for r in range(max_len):
        ws2.append([None] * (2 * len(group_cols)))
        excel_row = ws2.max_row

        for k, g in enumerate(group_cols):
            lemma, pct = per_group_sorted[g][r]
            lemma_col_idx = 1 + 2 * k
            pct_col_idx = 2 + 2 * k

            cell_lemma = ws2.cell(row=excel_row, column=lemma_col_idx, value=lemma)
            ws2.cell(row=excel_row, column=pct_col_idx, value=pct)

            if best_group_for_lemma.get(lemma) == g:
                cell_lemma.font = bold_font

    for col in range(2, 2 * len(group_cols) + 1, 2):
        for i in range(2, ws2.max_row + 1):
            ws2.cell(row=i, column=col).number_format = "0.00%"

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(ws2.max_column)}{ws2.max_row}"

    wb.save(xlsx_path)


# ------------------------- main -------------------------

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("input_csv", help="MLM output CSV with group columns appended")
    ap.add_argument("output", help="Output filename (.csv or .xlsx)")
    ap.add_argument("--lemma-col", default="lemma", help="Lemma column name (default: lemma)")
    ap.add_argument(
        "--group-cols",
        nargs="+",
        default=None,
        help="Optional explicit group column names; otherwise inferred from header",
    )
    ap.add_argument(
        "--second-threshold",
        type=float,
        default=0.50,
        help="Color lemma blue if 2nd-best >= threshold * best (default: 0.50)",
    )
    ap.add_argument("--encoding", default="utf-8")
    args = ap.parse_args()

    if not (0.0 <= args.second_threshold <= 1.0):
        raise SystemExit("--second-threshold must be between 0 and 1.")

    out_ext = os.path.splitext(args.output)[1].lower()
    if out_ext not in {".csv", ".xlsx"}:
        raise SystemExit("Output filename must end with .csv or .xlsx")

    # ---------- read + aggregate ----------
    with open(args.input_csv, newline="", encoding=args.encoding) as fin:
        reader = csv.DictReader(fin)
        if reader.fieldnames is None:
            raise SystemExit("Input CSV has no header.")

        if args.lemma_col not in reader.fieldnames:
            raise SystemExit(f"Missing lemma column {args.lemma_col!r}")

        group_cols = args.group_cols or infer_group_cols(reader.fieldnames)
        if not group_cols:
            raise SystemExit("Could not infer group columns. Provide --group-cols explicitly.")

        sums: Dict[str, Dict[str, float]] = {g: defaultdict(float) for g in group_cols}
        counts: Dict[str, int] = defaultdict(int)

        for row in reader:
            lemma = (row.get(args.lemma_col) or "").strip()
            if not lemma:
                continue

            counts[lemma] += 1
            for g in group_cols:
                val = (row.get(g) or "").strip()
                if not val:
                    continue
                try:
                    sums[g][lemma] += float(val)
                except ValueError:
                    pass

    means: Dict[str, Dict[str, float]] = {
        g: {lemma: (sums[g][lemma] / counts[lemma]) for lemma in counts}
        for g in group_cols
    }

    best_group_for_lemma: Dict[str, str] = {}
    for lemma in counts:
        best_group_for_lemma[lemma] = max(group_cols, key=lambda g: means[g][lemma])

    # ---------- write output ----------
    if out_ext == ".csv":
        with open(args.output, "w", newline="", encoding=args.encoding) as fout:
            writer = csv.DictWriter(
                fout,
                fieldnames=[args.lemma_col] + group_cols + ["count"],
            )
            writer.writeheader()
            for lemma in sorted(counts):
                row = {args.lemma_col: lemma, "count": counts[lemma]}
                for g in group_cols:
                    row[g] = f"{means[g][lemma]:.10g}"
                writer.writerow(row)
    else:
        write_excel(
            xlsx_path=args.output,
            lemma_col=args.lemma_col,
            group_cols=group_cols,
            counts=counts,
            means=means,
            best_group_for_lemma=best_group_for_lemma,
            second_threshold=args.second_threshold,
        )


if __name__ == "__main__":
    main()