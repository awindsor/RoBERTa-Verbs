#!/usr/bin/env python3
"""
MLM Group Aggregator - Recompute group probabilities from MLM predictions.

Supports both CLI and GUI modes:
  - Run without arguments: launches GUI (requires PySide6)
  - Run with arguments: uses CLI mode

Use-case:
  - You already ran RoBERTaMaskedLanguageModelVerbs.py and have a CSV with token_i/prob_i
  - You have a grouping CSV (columns=groups, cells=lemmas; header row=group names)
  - You want to compute per-row group probabilities by summing probs of tokens in each group
  - Can load settings from RoBERTaMaskedLanguageModelVerbs.json or MLMGroupAggregator.json

Output modes:
  1) Default: write ALL original columns + appended group columns
  2) --short: write only lemma + group columns (optionally include count)

Requirements:
  pip install lemminflect
  pip install PySide6  # for GUI mode only
    pip install openpyxl  # for XLSX output (optional)

CLI Examples:
  python MLMGroupAggregator.py mlm_out.csv groups.csv output.csv
  python MLMGroupAggregator.py mlm_out.csv groups.csv output.csv --short
  python MLMGroupAggregator.py mlm_out.csv groups.csv output.csv --load-metadata mlm_out.json

GUI Mode:
  python MLMGroupAggregator.py
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import os
import re
import sys
import time
from collections import deque
from concurrent.futures import ProcessPoolExecutor, TimeoutError as FuturesTimeoutError
from datetime import datetime, timezone
from functools import lru_cache
from itertools import islice
from pathlib import Path
from typing import Any, Dict, Iterable, List, Set, Tuple, Optional, Sequence

from lemminflect import getLemma
import hashlib


PROB_COL_RE = re.compile(r"^prob_(\d+)$")

_WORKER_GROUP_LABELS: List[str] = []
_WORKER_LEMMA_TO_GROUPS: Dict[str, Set[str]] = {}
_WORKER_K = 0
_WORKER_LEMMA_COL = "lemma"
_WORKER_SHORT = False
_WORKER_GROUP_COLNAME_MAP: Dict[str, str] = {}


def default_worker_count(reserve_cores: int = 2) -> int:
    """Return a responsive default worker count, leaving some cores free."""
    cpu_total = os.cpu_count() or 1
    return max(1, cpu_total - reserve_cores)


# ============================================================================
# METADATA FUNCTIONS
# ============================================================================

def get_aggregator_version_info() -> Dict:
    """Get version information for MLMGroupAggregator dependencies."""
    try:
        import lemminflect
        lemminflect_version = lemminflect.__version__
    except (ImportError, AttributeError):
        lemminflect_version = "unknown"
    
    return {
        "lemminflect": lemminflect_version,
    }


def compute_file_md5(path: Path) -> str:
    """Compute MD5 checksum of a file."""
    md5 = hashlib.md5()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            md5.update(chunk)
    return md5.hexdigest()


def save_aggregation_metadata(
    output_path: Path,
    mlm_csv_path: Path,
    group_csv_path: Path,
    mlm_checksum: str,
    group_checksum: str,
    output_checksum: str,
    settings: Dict[str, Any],
    stats: Optional[Dict[str, Any]],
    group_labels: List[str],
    lemma_to_groups: Dict[str, Set[str]],
    source_metadata: Optional[Dict] = None,
    command: Optional[str] = None,
    status: str = "completed",
) -> None:
    """Save aggregation metadata to JSON file alongside output.
    
    Args:
        source_metadata: Metadata from previous tool (e.g., RoBERTaMaskedLanguageModelVerbs) for chaining
    """
    json_path = output_path.with_suffix(".json")
    
    # Restructure lemma_to_groups back into the original file format:
    # groups as keys, lemmas as sorted lists (matching the input CSV structure)
    groups_structure = {}
    for group_name in group_labels:
        lemmas_in_group = sorted([
            lemma for lemma, groups in lemma_to_groups.items()
            if group_name in groups
        ])
        groups_structure[group_name] = lemmas_in_group
    
    metadata = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "tool": "MLMGroupAggregator",
        "versions": get_aggregator_version_info(),
        "status": status,
        "output_file": str(output_path),
        "output_checksum": output_checksum,
        "input_files": {
            "mlm_csv": str(mlm_csv_path),
            "group_csv": str(group_csv_path),
        },
        "input_checksums": {
            "mlm_csv": mlm_checksum,
            "group_csv": group_checksum,
        },
        "command": command,
        "groups": groups_structure,
        "settings": {
            "top_k": settings.get("top_k"),
            "lemma_col": settings.get("lemma_col", "lemma"),
            "short": settings.get("short", False),
            "include_count": settings.get("include_count", False),
            "encoding": settings.get("encoding", "utf-8-sig"),
            "workers": settings.get("workers"),
            "chunk_size": settings.get("chunk_size"),
        },
        "statistics": stats,
    }
    
    # Include source metadata for pipeline traceability
    if source_metadata:
        metadata["source_metadata"] = source_metadata
    
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(metadata, f, indent=2)


def save_incomplete_aggregation_metadata(
    output_path: Path,
    mlm_csv_path: Path,
    group_csv_path: Path,
    settings: Dict[str, Any],
    group_labels: List[str],
    lemma_to_groups: Dict[str, Set[str]],
    source_metadata: Optional[Dict] = None,
    command: Optional[str] = None,
) -> None:
    """Write restartable metadata before aggregation begins."""
    save_aggregation_metadata(
        output_path=output_path,
        mlm_csv_path=mlm_csv_path,
        group_csv_path=group_csv_path,
        mlm_checksum="",
        group_checksum="",
        output_checksum="",
        settings=settings,
        stats=None,
        group_labels=group_labels,
        lemma_to_groups=lemma_to_groups,
        source_metadata=source_metadata,
        command=command,
        status="incomplete",
    )


def load_aggregation_metadata(json_path: Path) -> Any:
    """Load aggregation metadata from JSON file."""
    with json_path.open("r", encoding="utf-8") as f:
        return json.load(f)


def load_mlm_metadata(json_path: Path) -> Any:
    """Load MLM metadata from RoBERTaMaskedLanguageModelVerbs JSON."""
    with json_path.open("r", encoding="utf-8") as f:
        return json.load(f)


def verify_input_checksums(
    mlm_path: Path,
    group_path: Path,
    metadata: Dict,
) -> Tuple[Dict[str, bool], str]:
    """
    Verify input file checksums against metadata.
    
    Returns:
        (status_dict, message) where status_dict has keys 'mlm_ok', 'group_ok'
    """
    status = {"mlm_ok": True, "group_ok": True}
    messages = []
    
    input_checksums = metadata.get("input_checksums", {})
    
    # Check MLM CSV
    if "mlm_csv" in input_checksums:
        expected = input_checksums["mlm_csv"]
        if mlm_path.exists():
            actual = compute_file_md5(mlm_path)
            if actual != expected:
                status["mlm_ok"] = False
                messages.append("⚠ MLM CSV has changed (checksum mismatch)")
        else:
            status["mlm_ok"] = False
            messages.append("⚠ MLM CSV not found")
    
    # Check group CSV
    if "group_csv" in input_checksums:
        expected = input_checksums["group_csv"]
        if group_path.exists():
            actual = compute_file_md5(group_path)
            if actual != expected:
                status["group_ok"] = False
                messages.append("⚠ Group CSV has changed (checksum mismatch)")
        else:
            status["group_ok"] = False
            messages.append("⚠ Group CSV not found")
    
    message = " | ".join(messages) if messages else "✓ All input files verified"
    return status, message


def _write_groups_sheet_openpyxl(wb, group_labels: List[str], lemma_to_groups: Dict[str, Set[str]]) -> None:
    """
    Add a sheet named "Groups" to the provided openpyxl workbook with
    columns for each group label and rows listing lemmas belonging to
    that group. Lemmas are sorted alphabetically per group.
    """
    ws = wb.create_sheet("Groups")
    ws.append(group_labels)
    group_to_lemmas: Dict[str, List[str]] = {g: [] for g in group_labels}
    for lemma, groups in lemma_to_groups.items():
        for g in groups:
            if g in group_to_lemmas:
                group_to_lemmas[g].append(lemma)
    for g in group_labels:
        group_to_lemmas[g].sort()
    max_len = max((len(v) for v in group_to_lemmas.values()), default=0)
    for i in range(max_len):
        row = [group_to_lemmas[g][i] if i < len(group_to_lemmas[g]) else "" for g in group_labels]
        ws.append(row)


# ============================================================================
# CORE LOGIC (shared by CLI and GUI)
# ============================================================================


@lru_cache(maxsize=200_000)
def normalize_pred_token(tok: str) -> str:
    """Normalize an MLM prediction token to a lemma key used by groups.

    - Strips common subword markers (e.g., 'Ġ', '▁').
    - Lowercases the token.
    - Lemmatizes as a verb using lemminflect's getLemma.
    """
    if tok is None:
        return ""
    s = tok.strip()
    if not s:
        return ""
    # Strip common RoBERTa/SentencePiece markers
    if s.startswith("Ġ"):
        s = s[1:]
    if s.startswith("▁"):
        s = s[1:]
    s = s.strip().lower()
    if not s:
        return ""
    try:
        lemmas = getLemma(s, "VERB")
        if lemmas:
            return (lemmas[0] or s).lower()
    except Exception:
        pass
    return s


def _init_aggregation_worker(
    group_labels: List[str],
    lemma_to_groups: Dict[str, Set[str]],
    k: int,
    lemma_col: str,
    short: bool,
    group_colname_map: Dict[str, str],
) -> None:
    """Initialize per-process state for batch aggregation workers."""
    global _WORKER_GROUP_LABELS, _WORKER_LEMMA_TO_GROUPS, _WORKER_K
    global _WORKER_LEMMA_COL, _WORKER_SHORT, _WORKER_GROUP_COLNAME_MAP
    _WORKER_GROUP_LABELS = group_labels
    _WORKER_LEMMA_TO_GROUPS = lemma_to_groups
    _WORKER_K = k
    _WORKER_LEMMA_COL = lemma_col
    _WORKER_SHORT = short
    _WORKER_GROUP_COLNAME_MAP = group_colname_map


def _process_aggregation_batch(rows: List[Dict[str, str]]) -> List[Dict[str, Any]]:
    """Aggregate one batch of MLM rows. Output order matches input order."""
    results: List[Dict[str, Any]] = []
    group_labels = _WORKER_GROUP_LABELS
    lemma_to_groups = _WORKER_LEMMA_TO_GROUPS
    k = _WORKER_K
    lemma_col = _WORKER_LEMMA_COL
    short = _WORKER_SHORT
    group_colname_map = _WORKER_GROUP_COLNAME_MAP

    for row in rows:
        try:
            lemma = (row.get(lemma_col) or "").strip()
            if not lemma:
                raise ValueError(f"Empty lemma in column {lemma_col!r}")

            group_sums: Dict[str, float] = {g: 0.0 for g in group_labels}

            for i in range(1, k + 1):
                tok = row.get(f"token_{i}", "")
                p_str = row.get(f"prob_{i}", "")
                if not tok or not p_str:
                    continue
                try:
                    prob = float(p_str)
                except ValueError:
                    continue

                key = normalize_pred_token(tok)
                gs = lemma_to_groups.get(key)
                if not gs:
                    continue
                for g in gs:
                    group_sums[g] += prob

            if short:
                out_row = {lemma_col: lemma}
                for g in group_labels:
                    out_row[group_colname_map[g]] = f"{group_sums[g]:.10g}"
            else:
                out_row = dict(row)
                for g in group_labels:
                    out_row[group_colname_map[g]] = f"{group_sums[g]:.10g}"

            results.append({"ok": True, "lemma": lemma, "out_row": out_row})
        except Exception as e:
            results.append({"ok": False, "error": str(e)})

    return results


def _iter_row_batches(reader: csv.DictReader, chunk_size: int) -> Iterable[List[Dict[str, str]]]:
    """Yield fixed-size batches of rows from a DictReader."""
    while True:
        batch = list(islice(reader, chunk_size))
        if not batch:
            return
        yield batch


def _iter_row_batches_with_offsets(
    reader: csv.DictReader,
    file_obj,
    chunk_size: int,
) -> Iterable[Tuple[List[Dict[str, str]], int]]:
    """Yield fixed-size batches of rows plus the current input file byte offset."""
    while True:
        batch = list(islice(reader, chunk_size))
        if not batch:
            return
        yield batch, file_obj.buffer.tell()


def _terminate_process_pool(executor: ProcessPoolExecutor, progress_callback=None) -> None:
    """Aggressively stop a process pool, including currently running workers."""
    if progress_callback:
        progress_callback("Terminating worker processes...")
    executor.shutdown(wait=False, cancel_futures=True)
    processes = getattr(executor, "_processes", {}) or {}
    for proc in processes.values():
        try:
            if proc.is_alive():
                proc.terminate()
        except Exception:
            pass
    for proc in processes.values():
        try:
            proc.join(timeout=0.2)
            if proc.is_alive() and hasattr(proc, "kill"):
                proc.kill()
        except Exception:
            pass


def _get_future_result_with_cancellation(
    future,
    executor: ProcessPoolExecutor,
    is_cancelled,
    progress_callback=None,
):
    """Poll a future so cancellation can tear down running workers promptly."""
    while True:
        if is_cancelled():
            if progress_callback:
                progress_callback("Cancellation requested. Stopping running batches...")
            _terminate_process_pool(executor, progress_callback)
            raise KeyboardInterrupt("User cancelled")
        try:
            return future.result(timeout=0.2)
        except FuturesTimeoutError:
            continue


def _format_bytes(num_bytes: int) -> str:
    """Render a byte count with binary units for the progress label."""
    value = float(max(0, num_bytes))
    for unit in ["B", "KB", "MB", "GB", "TB"]:
        if value < 1024.0 or unit == "TB":
            if unit == "B":
                return f"{int(value)} {unit}"
            return f"{value:.1f} {unit}"
        value /= 1024.0


def load_groups_csv(
    group_csv_path: str,
    encoding: str,
    logger: logging.Logger,
    progress_callback=None,
) -> Tuple[List[str], Dict[str, Set[str]]]:
    """Load groups from a wide CSV (columns=groups, rows=list lemmas) into mapping.

    Returns:
        (group_labels, lemma_to_groups)
    """
    p = Path(group_csv_path)
    if not p.exists():
        raise FileNotFoundError(f"Group CSV not found: {p}")
    logger.info(f"Opening group CSV: {p}")
    if progress_callback:
        progress_callback(f"Opening group CSV: {p}")
    with p.open("r", newline="", encoding=encoding) as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if not header:
            raise ValueError("Group CSV is empty or missing header row")
        group_labels: List[str] = [h.strip() for h in header if h and h.strip()]
        logger.info(f"Found {len(group_labels)} groups in group CSV header")
        if progress_callback:
            progress_callback(f"Found {len(group_labels)} groups in group CSV header")
        lemma_to_groups: Dict[str, Set[str]] = {}
        inflection_logged = False
        for row in reader:
            for idx, g in enumerate(group_labels):
                if idx >= len(row):
                    continue
                cell = (row[idx] or "").strip()
                if not cell:
                    continue
                if not inflection_logged:
                    logger.info("Inflecting group lemmas with lemminflect during group normalization...")
                    if progress_callback:
                        progress_callback("Inflecting group lemmas with lemminflect during group normalization...")
                    inflection_logged = True
                key = normalize_pred_token(cell)
                if not key:
                    continue
                s = lemma_to_groups.get(key)
                if s is None:
                    s = set()
                    lemma_to_groups[key] = s
                s.add(g)
    logger.info(f"Loaded {len(group_labels)} groups; {len(lemma_to_groups)} unique lemmas mapped")
    return group_labels, lemma_to_groups


def infer_top_k(fieldnames: Sequence[str]) -> int:
    """Infer top_k from fieldnames by counting prob_i columns."""
    max_k = 0
    for name in fieldnames:
        m = PROB_COL_RE.match(name)
        if m:
            try:
                idx = int(m.group(1))
                if idx > max_k:
                    max_k = idx
            except Exception:
                continue
    return max_k


def build_group_output_columns(fieldnames: Sequence[str], group_labels: List[str]) -> Tuple[List[str], Dict[str, str]]:
    """Create safe output column names for each group and a mapping from group->column.

    Ensures names don't collide with existing fieldnames by appending numeric suffixes if needed.
    """
    existing = set(fieldnames)
    out_cols: List[str] = []
    col_map: Dict[str, str] = {}
    for g in group_labels:
        base = f"group_prob_{g}"
        name = base
        suffix = 1
        while name in existing or name in out_cols:
            name = f"{base}_{suffix}"
            suffix += 1
        out_cols.append(name)
        col_map[g] = name
    return out_cols, col_map

# ============================================================================
# CLI MODE
# ============================================================================

def run_cli() -> None:
    """Run in CLI mode."""
    ap = argparse.ArgumentParser(
        description="MLM Group Aggregator - Recompute group probabilities from MLM predictions"
    )
    ap.add_argument("mlm_csv", help="MLM output CSV with token_i/prob_i columns")
    ap.add_argument("group_csv", help="Group CSV (columns=groups, cells=lemmas)")
    ap.add_argument("output_csv", help="Output CSV with group probabilities appended")
    ap.add_argument("--top-k", type=int, default=0, help="Use this top_k instead of inferring")
    ap.add_argument("--lemma-col", default="lemma", help="Lemma column name (default: lemma)")
    ap.add_argument("--short", action="store_true", help="Output only lemma + group columns")
    ap.add_argument("--include-count", action="store_true", help="With --short, include per-lemma count")
    ap.add_argument("--encoding", default="utf-8-sig", help="CSV encoding (default: utf-8-sig)")
    ap.add_argument("--log-every", type=int, default=100000, help="Log progress every N rows")
    ap.add_argument("--log-level", default="INFO", help="Logging level (default: INFO)")
    ap.add_argument(
        "--workers",
        type=int,
        default=None,
        help="Number of worker processes to use (default: CPU cores minus 2; use 1 to disable)",
    )
    ap.add_argument(
        "--chunk-size",
        type=int,
        default=None,
        help="Rows per worker batch (default: 2000)",
    )
    ap.add_argument(
        "--load-metadata",
        help="Load settings from metadata JSON (RoBERTaMaskedLanguageModelVerbs.json or MLMGroupAggregator.json)"
    )
    args = ap.parse_args()
    
    logging.basicConfig(
        level=getattr(logging, args.log_level.upper(), logging.INFO),
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )
    logger = logging.getLogger("aggregate")
    
    # Load metadata if provided
    metadata = None
    source_metadata = None  # Store the source metadata for chaining
    
    if args.load_metadata:
        metadata_path = Path(args.load_metadata)
        if not metadata_path.exists():
            raise SystemExit(f"Metadata file not found: {metadata_path}")
        
        with metadata_path.open("r", encoding="utf-8") as f:
            try:
                raw_metadata = json.load(f)
            except json.JSONDecodeError as e:
                raise SystemExit(f"Failed to parse JSON metadata: {e}")
        
        # Detect what type of metadata this is
        tool_name = raw_metadata.get("tool", "unknown")
        logger.info(f"This metadata is from '{tool_name}'")
        
        # Validate tool type
        supported_tools = ["RoBERTaMaskedLanguageModelVerbs", "MLMGroupAggregator"]
        if tool_name not in supported_tools:
            raise SystemExit(
                f"Unsupported metadata source: '{tool_name}'\n"
                "MLMGroupAggregator accepts metadata from:\n"
                "  - RoBERTaMaskedLanguageModelVerbs\n"
                "  - MLMGroupAggregator\n"
                "Received: {tool_name}"
            )
        
        metadata = raw_metadata
        source_metadata = raw_metadata.get("source_metadata")
        
        # Extract and apply settings from loaded metadata
        settings = metadata.get("settings", {})
        
        # Apply settings from metadata
        if not args.top_k or args.top_k <= 0:
            args.top_k = settings.get("top_k", 0)
        if args.lemma_col == "lemma":
            args.lemma_col = settings.get("lemma_col", "lemma")
        if not args.short:
            args.short = settings.get("short", False)
        if not args.include_count:
            args.include_count = settings.get("include_count", False)
        if args.workers is None:
            args.workers = settings.get("workers")
        if args.chunk_size is None:
            args.chunk_size = settings.get("chunk_size")
        
        # Infer input file paths based on tool type
        if tool_name == "RoBERTaMaskedLanguageModelVerbs":
            if args.mlm_csv is None:
                args.mlm_csv = metadata.get("output_file")
        elif tool_name == "MLMGroupAggregator":
            input_files = metadata.get("input_files", {})
            if args.mlm_csv is None:
                args.mlm_csv = input_files.get("mlm_csv")
            if args.group_csv is None:
                args.group_csv = input_files.get("group_csv")
    
    mlm_path = Path(args.mlm_csv)
    group_path = Path(args.group_csv)
    output_path = Path(args.output_csv)
    
    if not mlm_path.exists():
        raise SystemExit(f"MLM CSV not found: {mlm_path}")
    if not group_path.exists():
        raise SystemExit(f"Group CSV not found: {group_path}")
    if args.workers is None:
        args.workers = default_worker_count()
    if args.chunk_size is None:
        args.chunk_size = 2000
    if args.workers < 1:
        raise SystemExit("--workers must be >= 1")
    if args.chunk_size < 1:
        raise SystemExit("--chunk-size must be >= 1")
    
    # Verify checksums if metadata was loaded
    if metadata:
        status, msg = verify_input_checksums(mlm_path, group_path, metadata)
        if not status["mlm_ok"] or not status["group_ok"]:
            logger.warning(msg)
    
    start_time = time.time()
    group_labels, lemma_to_groups = load_groups_csv(args.group_csv, "utf-8-sig", logger)
    aggregation_settings = {
        "top_k": args.top_k if args.top_k > 0 else None,
        "lemma_col": args.lemma_col,
        "short": args.short,
        "include_count": args.include_count,
        "encoding": args.encoding,
        "workers": args.workers,
        "chunk_size": args.chunk_size,
        "output_format": "xlsx" if output_path.suffix.lower() == ".xlsx" else "csv",
    }
    
    processed = 0
    skipped = 0
    lemma_counts: Dict[str, int] = {}
    
    try:
        with mlm_path.open(newline="", encoding=args.encoding) as fin:
            reader = csv.DictReader(fin)
            if reader.fieldnames is None:
                raise SystemExit("Input MLM CSV has no header.")

            fieldnames = reader.fieldnames

            if args.lemma_col not in fieldnames:
                raise SystemExit(f"Input MLM CSV missing lemma column {args.lemma_col!r}")

            k = args.top_k if args.top_k > 0 else infer_top_k(fieldnames)
            if k <= 0:
                raise SystemExit("Could not infer top_k from header; provide --top-k explicitly.")
            aggregation_settings["top_k"] = k

            # Verify required token/prob columns
            missing_cols = []
            for i in range(1, k + 1):
                if f"token_{i}" not in fieldnames:
                    missing_cols.append(f"token_{i}")
                if f"prob_{i}" not in fieldnames:
                    missing_cols.append(f"prob_{i}")
            if missing_cols:
                raise SystemExit(
                    f"Missing required token/prob columns for top_k={k}: "
                    f"{missing_cols[:10]}{' ...' if len(missing_cols) > 10 else ''}"
                )

            # Build output columns
            group_out_cols, group_colname_map = build_group_output_columns(fieldnames, group_labels)
            worker_count = args.workers
            use_parallel = worker_count > 1
            if use_parallel:
                logger.info(
                    "Using %d worker processes with chunk size %d",
                    worker_count,
                    args.chunk_size,
                )
            else:
                logger.info("Using single-process aggregation")

            if args.short:
                out_fieldnames = [args.lemma_col] + group_out_cols
                if args.include_count:
                    out_fieldnames.append("lemma_count")
            else:
                out_fieldnames = list(fieldnames) + group_out_cols

            is_xlsx = output_path.suffix.lower() == ".xlsx"

            save_incomplete_aggregation_metadata(
                output_path=output_path,
                mlm_csv_path=mlm_path,
                group_csv_path=group_path,
                settings=aggregation_settings,
                group_labels=group_labels,
                lemma_to_groups=lemma_to_groups,
                source_metadata=source_metadata,
                command=" ".join(sys.argv),
            )
            logger.info(f"✓ Saved incomplete metadata to {output_path.with_suffix('.json')}")

            if is_xlsx:
                try:
                    from openpyxl import Workbook
                    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
                except ImportError:
                    raise SystemExit(
                        "XLSX output requested but openpyxl is not installed. Install with: pip install openpyxl"
                    )

                def _clean_cell(val):
                    # Strip characters Excel cannot store (e.g., ASCII control chars)
                    if isinstance(val, str):
                        return ILLEGAL_CHARACTERS_RE.sub("", val)
                    return val

                # Write-only workbook keeps memory usage low for large files
                wb = Workbook(write_only=True)
                ws = wb.create_sheet("Aggregated")
                ws.append([_clean_cell(c) for c in out_fieldnames])

                row_batches = _iter_row_batches(reader, args.chunk_size)
                if use_parallel:
                    with ProcessPoolExecutor(
                        max_workers=worker_count,
                        initializer=_init_aggregation_worker,
                        initargs=(group_labels, lemma_to_groups, k, args.lemma_col, args.short, group_colname_map),
                    ) as executor:
                        result_batches = executor.map(_process_aggregation_batch, row_batches)
                        for batch_results in result_batches:
                            for result in batch_results:
                                if result["ok"]:
                                    out_row = result["out_row"]
                                    if args.short and args.include_count:
                                        lemma = result["lemma"]
                                        lemma_counts[lemma] = lemma_counts.get(lemma, 0) + 1
                                        out_row["lemma_count"] = str(lemma_counts[lemma])
                                    ws.append([_clean_cell(out_row.get(col, "")) for col in out_fieldnames])
                                    processed += 1
                                else:
                                    skipped += 1
                                    if skipped <= 10:
                                        logger.warning(f"Skipping row due to error: {result['error']}")
                            if args.log_every > 0 and processed and processed % args.log_every < len(batch_results):
                                logger.info(f"Written: {processed:,} | skipped: {skipped:,}")
                else:
                    _init_aggregation_worker(group_labels, lemma_to_groups, k, args.lemma_col, args.short, group_colname_map)
                    for batch in row_batches:
                        for result in _process_aggregation_batch(batch):
                            if result["ok"]:
                                out_row = result["out_row"]
                                if args.short and args.include_count:
                                    lemma = result["lemma"]
                                    lemma_counts[lemma] = lemma_counts.get(lemma, 0) + 1
                                    out_row["lemma_count"] = str(lemma_counts[lemma])
                                ws.append([_clean_cell(out_row.get(col, "")) for col in out_fieldnames])
                                processed += 1
                            else:
                                skipped += 1
                                if skipped <= 10:
                                    logger.warning(f"Skipping row due to error: {result['error']}")
                        if args.log_every > 0 and processed and processed % args.log_every < len(batch):
                            logger.info(f"Written: {processed:,} | skipped: {skipped:,}")

                # Add groups sheet and save workbook
                _write_groups_sheet_openpyxl(wb, group_labels, lemma_to_groups)
                wb.save(str(output_path))

            else:
                with output_path.open("w", newline="", encoding=args.encoding) as fout:
                    writer = csv.DictWriter(fout, fieldnames=out_fieldnames)
                    writer.writeheader()
                    row_batches = _iter_row_batches(reader, args.chunk_size)
                    if use_parallel:
                        with ProcessPoolExecutor(
                            max_workers=worker_count,
                            initializer=_init_aggregation_worker,
                            initargs=(group_labels, lemma_to_groups, k, args.lemma_col, args.short, group_colname_map),
                        ) as executor:
                            result_batches = executor.map(_process_aggregation_batch, row_batches)
                            for batch_results in result_batches:
                                for result in batch_results:
                                    if result["ok"]:
                                        out_row = result["out_row"]
                                        if args.short and args.include_count:
                                            lemma = result["lemma"]
                                            lemma_counts[lemma] = lemma_counts.get(lemma, 0) + 1
                                            out_row["lemma_count"] = str(lemma_counts[lemma])
                                        writer.writerow(out_row)
                                        processed += 1
                                    else:
                                        skipped += 1
                                        if skipped <= 10:
                                            logger.warning(f"Skipping row due to error: {result['error']}")
                                if args.log_every > 0 and processed and processed % args.log_every < len(batch_results):
                                    logger.info(f"Written: {processed:,} | skipped: {skipped:,}")
                    else:
                        _init_aggregation_worker(group_labels, lemma_to_groups, k, args.lemma_col, args.short, group_colname_map)
                        for batch in row_batches:
                            for result in _process_aggregation_batch(batch):
                                if result["ok"]:
                                    out_row = result["out_row"]
                                    if args.short and args.include_count:
                                        lemma = result["lemma"]
                                        lemma_counts[lemma] = lemma_counts.get(lemma, 0) + 1
                                        out_row["lemma_count"] = str(lemma_counts[lemma])
                                    writer.writerow(out_row)
                                    processed += 1
                                else:
                                    skipped += 1
                                    if skipped <= 10:
                                        logger.warning(f"Skipping row due to error: {result['error']}")
                            if args.log_every > 0 and processed and processed % args.log_every < len(batch):
                                logger.info(f"Written: {processed:,} | skipped: {skipped:,}")
        
        elapsed_time = time.time() - start_time
        
        # Compute checksums and save metadata
        logger.info("Computing file checksums...")
        mlm_checksum = compute_file_md5(mlm_path)
        group_checksum = compute_file_md5(group_path)
        output_checksum = compute_file_md5(output_path)
        
        stats = {
            "rows_written": processed,
            "rows_skipped": skipped,
            "groups_created": len(group_labels),
            "elapsed_seconds": round(elapsed_time, 2),
        }
        
        save_aggregation_metadata(
            output_path,
            mlm_path,
            group_path,
            mlm_checksum,
            group_checksum,
            output_checksum,
            aggregation_settings,
            stats,
            group_labels,
            lemma_to_groups,
            source_metadata,
            " ".join(sys.argv),
        )
        
        logger.info(f"Done. Written={processed:,} skipped={skipped:,} output={args.output_csv}")
        logger.info(f"✓ Saved metadata to {output_path.with_suffix('.json')}")
    
    except KeyboardInterrupt:
        logger.info("\\nCancellation requested. Saving incomplete results...")
        elapsed_time = time.time() - start_time
        
        # Save metadata with incomplete flag
        mlm_checksum = compute_file_md5(mlm_path)
        group_checksum = compute_file_md5(group_path)
        output_checksum = compute_file_md5(output_path) if output_path.exists() else ""
        
        stats = {
            "rows_written": processed,
            "rows_skipped": skipped,
            "groups_created": len(group_labels),
            "elapsed_seconds": round(elapsed_time, 2),
            "incomplete": True,
        }
        
        save_aggregation_metadata(
            output_path,
            mlm_path,
            group_path,
            mlm_checksum,
            group_checksum,
            output_checksum,
            aggregation_settings,
            stats,
            group_labels,
            lemma_to_groups,
            source_metadata,
            " ".join(sys.argv),
            "stopped_by_user",
        )
        
        logger.info(f"Cancelled. Written={processed:,} skipped={skipped:,}")
        logger.info(f"✓ Saved incomplete metadata to {output_path.with_suffix('.json')}")
        raise SystemExit(0)


# ============================================================================
# GUI MODE
# ============================================================================

def run_gui() -> None:
    """Launch the GUI application."""
    try:
        from PySide6.QtCore import Qt, QThread, Signal, QSize
        from PySide6.QtWidgets import (
            QApplication,
            QMainWindow,
            QWidget,
            QVBoxLayout,
            QHBoxLayout,
            QGroupBox,
            QLabel,
            QLineEdit,
            QPushButton,
            QFileDialog,
            QTextEdit,
            QProgressBar,
            QCheckBox,
            QSpinBox,
            QMessageBox,
        )
    except ImportError:
        print("Error: PySide6 is required for GUI mode.")
        print("Install it with: pip install PySide6")
        print("\nOr run in CLI mode by providing arguments. Use --help for usage.")
        sys.exit(1)
    
    class AggregationWorker(QThread):
        """Worker thread for group aggregation without blocking UI."""
        
        progress_update = Signal(str)
        progress_value = Signal(int, int)  # processed, total
        finished = Signal(bool, str)
        
        def __init__(
            self,
            mlm_csv_path: Path,
            group_csv_path: Path,
            output_csv_path: Path,
            top_k: int,
            lemma_col: str,
            short: bool,
            include_count: bool,
            encoding: str,
            worker_count: int,
        ):
            super().__init__()
            self.mlm_csv_path = mlm_csv_path
            self.group_csv_path = group_csv_path
            self.output_csv_path = output_csv_path
            self.top_k = top_k
            self.lemma_col = lemma_col
            self.short = short
            self.include_count = include_count
            self.encoding = encoding
            self.worker_count = worker_count
            self.chunk_size = 2000
            self._cancelled = False
        
        def cancel(self):
            """Request cancellation of aggregation."""
            self._cancelled = True
        
        def run(self):
            """Run the aggregation in the worker thread."""
            try:
                start_time = time.time()
                self.progress_update.emit(f"Loading groups from: {self.group_csv_path}")
                
                # Create a dummy logger
                logger = logging.getLogger("gui_aggregate")
                if not logger.handlers:
                    logger.addHandler(logging.NullHandler())
                
                group_labels, lemma_to_groups = load_groups_csv(
                    str(self.group_csv_path),
                    "utf-8-sig",
                    logger,
                    self.progress_update.emit,
                )
                
                self.progress_update.emit(f"Reading: {self.mlm_csv_path}")
                self.progress_update.emit("")

                total_bytes = self.mlm_csv_path.stat().st_size
                self.progress_update.emit(f"MLM CSV size: {_format_bytes(total_bytes)}")

                processed = 0
                skipped = 0
                lemma_counts: Dict[str, int] = {}

                self.progress_update.emit("Opening MLM CSV and reading header...")
                with self.mlm_csv_path.open(newline="", encoding=self.encoding) as fin:
                    reader = csv.DictReader(fin)
                    if reader.fieldnames is None:
                        raise ValueError("Input MLM CSV has no header.")
                    fieldnames = reader.fieldnames

                    if self.lemma_col not in fieldnames:
                        raise ValueError(f"Missing lemma column: {self.lemma_col}")

                    self.progress_update.emit("Inferring top_k from MLM CSV header...")
                    k = self.top_k if self.top_k > 0 else infer_top_k(fieldnames)
                    if k <= 0:
                        raise ValueError("Could not infer top_k; specify --top-k")
                    self.progress_update.emit(f"Using top_k={k}")
                    aggregation_settings = {
                        "top_k": k,
                        "lemma_col": self.lemma_col,
                        "short": self.short,
                        "include_count": self.include_count,
                        "encoding": self.encoding,
                        "workers": self.worker_count,
                        "chunk_size": self.chunk_size,
                        "output_format": "xlsx" if self.output_csv_path.suffix.lower() == ".xlsx" else "csv",
                    }

                    self.progress_update.emit("Building output group columns...")
                    group_out_cols, group_colname_map = build_group_output_columns(fieldnames, group_labels)
                    self.progress_update.emit(
                        f"Using {self.worker_count} worker processes (chunk size {self.chunk_size})"
                    )

                    if self.short:
                        out_fieldnames = [self.lemma_col] + group_out_cols
                        if self.include_count:
                            out_fieldnames.append("lemma_count")
                    else:
                        out_fieldnames = list(fieldnames) + group_out_cols

                    is_xlsx = self.output_csv_path.suffix.lower() == ".xlsx"
                    save_incomplete_aggregation_metadata(
                        output_path=self.output_csv_path,
                        mlm_csv_path=self.mlm_csv_path,
                        group_csv_path=self.group_csv_path,
                        settings=aggregation_settings,
                        group_labels=group_labels,
                        lemma_to_groups=lemma_to_groups,
                    )
                    self.progress_update.emit(
                        f"Saved metadata {self.output_csv_path.with_suffix('.json').name}"
                    )

                    if is_xlsx:
                        try:
                            from openpyxl import Workbook
                            from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
                        except ImportError:
                            raise ValueError(
                                "XLSX output requested but openpyxl is not installed. Install with: pip install openpyxl"
                            )

                        def _clean_cell(val):
                            # Strip characters Excel cannot store (e.g., ASCII control chars)
                            if isinstance(val, str):
                                return ILLEGAL_CHARACTERS_RE.sub("", val)
                            return val

                        # Write-only workbook keeps memory usage low for large files
                        self.progress_update.emit("Initializing XLSX workbook...")
                        wb = Workbook(write_only=True)
                        ws = wb.create_sheet("Aggregated")
                        ws.append([_clean_cell(c) for c in out_fieldnames])
                        self.progress_update.emit("Preparing MLM row batches...")
                        row_batches = _iter_row_batches_with_offsets(reader, fin, self.chunk_size)
                        self.progress_update.emit("Resetting progress before batch loading...")
                        self.progress_value.emit(0, total_bytes)
                        self.progress_update.emit("Starting worker pool and queueing initial batches...")
                        with ProcessPoolExecutor(
                            max_workers=self.worker_count,
                            initializer=_init_aggregation_worker,
                            initargs=(group_labels, lemma_to_groups, k, self.lemma_col, self.short, group_colname_map),
                        ) as executor:
                            pending = deque()
                            max_pending = max(1, self.worker_count * 2)
                            for batch, bytes_read in row_batches:
                                pending.append((executor.submit(_process_aggregation_batch, batch), bytes_read))
                                if len(pending) < max_pending:
                                    continue
                                future, batch_bytes = pending.popleft()
                                batch_results = _get_future_result_with_cancellation(
                                    future,
                                    executor,
                                    lambda: self._cancelled,
                                    self.progress_update.emit,
                                )
                                for result in batch_results:
                                    if result["ok"]:
                                        out_row = result["out_row"]
                                        if self.short and self.include_count:
                                            lemma = result["lemma"]
                                            lemma_counts[lemma] = lemma_counts.get(lemma, 0) + 1
                                            out_row["lemma_count"] = str(lemma_counts[lemma])
                                        ws.append([_clean_cell(out_row.get(col, "")) for col in out_fieldnames])
                                        processed += 1
                                    else:
                                        skipped += 1
                                        if skipped <= 5:
                                            self.progress_update.emit(f"⚠ Skipping row: {result['error'][:50]}")
                                self.progress_value.emit(batch_bytes, total_bytes)
                                if self._cancelled:
                                    self.progress_update.emit("Cancellation requested. Stopping running batches...")
                                    _terminate_process_pool(executor, self.progress_update.emit)
                                    raise KeyboardInterrupt("User cancelled")
                            while pending:
                                future, batch_bytes = pending.popleft()
                                batch_results = _get_future_result_with_cancellation(
                                    future,
                                    executor,
                                    lambda: self._cancelled,
                                    self.progress_update.emit,
                                )
                                for result in batch_results:
                                    if result["ok"]:
                                        out_row = result["out_row"]
                                        if self.short and self.include_count:
                                            lemma = result["lemma"]
                                            lemma_counts[lemma] = lemma_counts.get(lemma, 0) + 1
                                            out_row["lemma_count"] = str(lemma_counts[lemma])
                                        ws.append([_clean_cell(out_row.get(col, "")) for col in out_fieldnames])
                                        processed += 1
                                    else:
                                        skipped += 1
                                        if skipped <= 5:
                                            self.progress_update.emit(f"⚠ Skipping row: {result['error'][:50]}")
                                self.progress_value.emit(batch_bytes, total_bytes)
                                if self._cancelled:
                                    self.progress_update.emit("Cancellation requested. Stopping running batches...")
                                    _terminate_process_pool(executor, self.progress_update.emit)
                                    raise KeyboardInterrupt("User cancelled")

                        _write_groups_sheet_openpyxl(wb, group_labels, lemma_to_groups)
                        wb.save(str(self.output_csv_path))

                    else:
                        self.progress_update.emit("Opening output CSV for writing...")
                        with self.output_csv_path.open("w", newline="", encoding=self.encoding) as fout:
                            writer = csv.DictWriter(fout, fieldnames=out_fieldnames)
                            writer.writeheader()
                            self.progress_update.emit("Preparing MLM row batches...")
                            row_batches = _iter_row_batches_with_offsets(reader, fin, self.chunk_size)
                            self.progress_update.emit("Resetting progress before batch loading...")
                            self.progress_value.emit(0, total_bytes)
                            self.progress_update.emit("Starting worker pool and queueing initial batches...")
                            with ProcessPoolExecutor(
                                max_workers=self.worker_count,
                                initializer=_init_aggregation_worker,
                                initargs=(group_labels, lemma_to_groups, k, self.lemma_col, self.short, group_colname_map),
                            ) as executor:
                                pending = deque()
                                max_pending = max(1, self.worker_count * 2)
                                for batch, bytes_read in row_batches:
                                    pending.append((executor.submit(_process_aggregation_batch, batch), bytes_read))
                                    if len(pending) < max_pending:
                                        continue
                                    future, batch_bytes = pending.popleft()
                                    batch_results = _get_future_result_with_cancellation(
                                        future,
                                        executor,
                                        lambda: self._cancelled,
                                        self.progress_update.emit,
                                    )
                                    for result in batch_results:
                                        if result["ok"]:
                                            out_row = result["out_row"]
                                            if self.short and self.include_count:
                                                lemma = result["lemma"]
                                                lemma_counts[lemma] = lemma_counts.get(lemma, 0) + 1
                                                out_row["lemma_count"] = str(lemma_counts[lemma])
                                            writer.writerow(out_row)
                                            processed += 1
                                        else:
                                            skipped += 1
                                            if skipped <= 5:
                                                self.progress_update.emit(f"⚠ Skipping row: {result['error'][:50]}")
                                    self.progress_value.emit(batch_bytes, total_bytes)
                                    if self._cancelled:
                                        self.progress_update.emit("Cancellation requested. Stopping running batches...")
                                        _terminate_process_pool(executor, self.progress_update.emit)
                                        raise KeyboardInterrupt("User cancelled")
                                while pending:
                                    future, batch_bytes = pending.popleft()
                                    batch_results = _get_future_result_with_cancellation(
                                        future,
                                        executor,
                                        lambda: self._cancelled,
                                        self.progress_update.emit,
                                    )
                                    for result in batch_results:
                                        if result["ok"]:
                                            out_row = result["out_row"]
                                            if self.short and self.include_count:
                                                lemma = result["lemma"]
                                                lemma_counts[lemma] = lemma_counts.get(lemma, 0) + 1
                                                out_row["lemma_count"] = str(lemma_counts[lemma])
                                            writer.writerow(out_row)
                                            processed += 1
                                        else:
                                            skipped += 1
                                            if skipped <= 5:
                                                self.progress_update.emit(f"⚠ Skipping row: {result['error'][:50]}")
                                    self.progress_value.emit(batch_bytes, total_bytes)
                                    if self._cancelled:
                                        self.progress_update.emit("Cancellation requested. Stopping running batches...")
                                        _terminate_process_pool(executor, self.progress_update.emit)
                                        raise KeyboardInterrupt("User cancelled")
                
                # Save metadata
                self.progress_update.emit("Computing checksums and saving metadata...")
                self.progress_value.emit(total_bytes, total_bytes)
                mlm_checksum = compute_file_md5(self.mlm_csv_path)
                group_checksum = compute_file_md5(self.group_csv_path)
                output_checksum = compute_file_md5(self.output_csv_path)
                
                stats = {
                    "rows_written": processed,
                    "rows_skipped": skipped,
                    "groups_created": len(group_labels),
                    "elapsed_seconds": round(time.time() - start_time, 2),
                }
                
                save_aggregation_metadata(
                    self.output_csv_path,
                    self.mlm_csv_path,
                    self.group_csv_path,
                    mlm_checksum,
                    group_checksum,
                    output_checksum,
                    aggregation_settings,
                    stats,
                    group_labels,
                    lemma_to_groups,
                    None,
                    None,
                )
                
                self.progress_update.emit(f"✓ Written {processed:,} rows ({skipped:,} skipped)")
                self.progress_update.emit(f"✓ Saved metadata: {self.output_csv_path.with_suffix('.json')}")
                self.finished.emit(True, f"Aggregation complete. Output: {self.output_csv_path}")
                
            except KeyboardInterrupt:
                self.progress_update.emit("Saving incomplete results...")
                # Save metadata with incomplete flag
                try:
                    mlm_checksum = compute_file_md5(self.mlm_csv_path)
                    group_checksum = compute_file_md5(self.group_csv_path)
                    output_checksum = compute_file_md5(self.output_csv_path) if self.output_csv_path.exists() else ""
                    
                    stats = {
                        "rows_written": processed,
                        "rows_skipped": skipped,
                        "groups_created": len(group_labels),
                        "elapsed_seconds": round(time.time() - start_time, 2),
                        "incomplete": True,
                    }
                    
                    save_aggregation_metadata(
                        self.output_csv_path,
                        self.mlm_csv_path,
                        self.group_csv_path,
                        mlm_checksum,
                        group_checksum,
                        output_checksum,
                        aggregation_settings,
                        stats,
                        group_labels,
                        lemma_to_groups,
                        None,
                        None,
                        "stopped_by_user",
                    )
                    self.progress_update.emit(
                        f"Saved metadata {self.output_csv_path.with_suffix('.json').name}"
                    )
                except Exception as meta_err:
                    self.progress_update.emit(f"⚠ Failed to save metadata: {meta_err}")
                
                self.finished.emit(False, f"Aggregation cancelled. Processed {processed:,} rows.")
                
            except Exception as e:
                self.progress_update.emit(f"✗ Error: {str(e)}")
                self.finished.emit(False, f"Error: {str(e)}")
    
    class MLMGroupAggregatorGUI(QMainWindow):
        """Main GUI window for MLM group aggregation."""
        
        def __init__(self):
            super().__init__()
            self.setWindowTitle("MLM Group Aggregator")
            self.setMinimumSize(QSize(900, 650))
            
            self.worker: Optional[AggregationWorker] = None
            self.init_ui()
        
        def init_ui(self):
            """Initialize the UI components."""
            central = QWidget()
            self.setCentralWidget(central)
            
            layout = QVBoxLayout(central)
            
            # Files section
            files_group = QGroupBox("Files")
            files_layout = QVBoxLayout()
            
            # MLM CSV
            mlm_layout = QHBoxLayout()
            mlm_layout.addWidget(QLabel("MLM CSV:"))
            self.mlm_input = QLineEdit()
            self.mlm_browse = QPushButton("Browse...")
            self.mlm_browse.clicked.connect(self.browse_mlm)
            mlm_layout.addWidget(self.mlm_input)
            mlm_layout.addWidget(self.mlm_browse)
            files_layout.addLayout(mlm_layout)
            
            # Group CSV
            group_layout = QHBoxLayout()
            group_layout.addWidget(QLabel("Group CSV:"))
            self.group_input = QLineEdit()
            self.group_browse = QPushButton("Browse...")
            self.group_browse.clicked.connect(self.browse_group)
            group_layout.addWidget(self.group_input)
            group_layout.addWidget(self.group_browse)
            files_layout.addLayout(group_layout)
            
            # Output File
            output_layout = QHBoxLayout()
            output_layout.addWidget(QLabel("Output File:"))
            self.output_input = QLineEdit("aggregated_output.csv")
            self.output_browse = QPushButton("Browse...")
            self.output_browse.clicked.connect(self.browse_output)
            output_layout.addWidget(self.output_input)
            output_layout.addWidget(self.output_browse)
            files_layout.addLayout(output_layout)
            
            files_group.setLayout(files_layout)
            layout.addWidget(files_group)
            
            # Settings section
            settings_group = QGroupBox("Settings")
            settings_layout = QVBoxLayout()
            
            # Load metadata button
            metadata_layout = QHBoxLayout()
            self.load_metadata_btn = QPushButton("Load Settings from JSON")
            self.load_metadata_btn.clicked.connect(self.load_metadata_dialog)
            metadata_layout.addWidget(self.load_metadata_btn)
            metadata_layout.addStretch()
            settings_layout.addLayout(metadata_layout)
            
            # Options
            options_layout = QHBoxLayout()
            
            top_k_layout = QVBoxLayout()
            top_k_layout.addWidget(QLabel("Top-K:"))
            self.topk_spin = QSpinBox()
            self.topk_spin.setMinimum(0)
            self.topk_spin.setMaximum(100)
            self.topk_spin.setValue(0)
            self.topk_spin.setToolTip("0 = infer from MLM CSV header")
            top_k_layout.addWidget(self.topk_spin)

            workers_layout = QVBoxLayout()
            workers_layout.addWidget(QLabel("Worker Processes:"))
            self.workers_spin = QSpinBox()
            self.workers_spin.setMinimum(1)
            self.workers_spin.setMaximum(max(1, os.cpu_count() or 1))
            self.workers_spin.setValue(default_worker_count())
            self.workers_spin.setToolTip("Default leaves 2 CPU cores free to keep the system responsive")
            workers_layout.addWidget(self.workers_spin)
            
            lemma_col_layout = QVBoxLayout()
            lemma_col_layout.addWidget(QLabel("Lemma Column:"))
            self.lemma_col_input = QLineEdit("lemma")
            lemma_col_layout.addWidget(self.lemma_col_input)
            
            options_layout.addLayout(top_k_layout)
            options_layout.addLayout(workers_layout)
            options_layout.addLayout(lemma_col_layout)
            options_layout.addStretch()
            settings_layout.addLayout(options_layout)
            
            # Checkboxes
            checkbox_layout = QHBoxLayout()
            self.short_check = QCheckBox("Short Mode (lemma + groups only)")
            self.count_check = QCheckBox("Include Lemma Count")
            self.count_check.setEnabled(False)
            self.short_check.toggled.connect(self.count_check.setEnabled)
            checkbox_layout.addWidget(self.short_check)
            checkbox_layout.addWidget(self.count_check)
            checkbox_layout.addStretch()
            settings_layout.addLayout(checkbox_layout)
            
            settings_group.setLayout(settings_layout)
            layout.addWidget(settings_group)
            
            # Progress section
            progress_group = QGroupBox("Progress")
            progress_layout = QVBoxLayout()
            self.progress_bar = QProgressBar()
            self.progress_bar.setMinimum(0)
            self.progress_bar.setMaximum(1)
            self.progress_bar.setValue(0)
            self.progress_bar.setTextVisible(True)
            progress_layout.addWidget(self.progress_bar)
            self.progress_label = QLabel("Ready")
            self.progress_label.setStyleSheet("color: #666; font-style: italic;")
            progress_layout.addWidget(self.progress_label)
            self.progress_text = QTextEdit()
            self.progress_text.setReadOnly(True)
            self.progress_text.setMaximumHeight(200)
            progress_layout.addWidget(self.progress_text)
            progress_group.setLayout(progress_layout)
            layout.addWidget(progress_group)
            
            # Control buttons
            button_layout = QHBoxLayout()
            self.start_btn = QPushButton("Start Aggregation")
            self.start_btn.clicked.connect(self.toggle_aggregation)
            self.start_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 8px;")
            button_layout.addStretch()
            button_layout.addWidget(self.start_btn)
            button_layout.addStretch()
            layout.addLayout(button_layout)
            
            central.setLayout(layout)
        
        def browse_mlm(self):
            file, _ = QFileDialog.getOpenFileName(self, "Select MLM CSV", "", "CSV Files (*.csv)")
            if file:
                self.mlm_input.setText(file)
                # Auto-suggest output in the same directory if not already set
                if not self.output_input.text() or self.output_input.text() == "aggregated_output.csv":
                    out_path = Path(file).with_name("aggregated_output.csv")
                    self.output_input.setText(str(out_path))
        
        def browse_group(self):
            file, _ = QFileDialog.getOpenFileName(self, "Select Group CSV", "", "CSV Files (*.csv)")
            if file:
                self.group_input.setText(file)
        
        def browse_output(self):
            file, _ = QFileDialog.getSaveFileName(
                self,
                "Select Output File",
                str(Path(self.mlm_input.text()).with_name("aggregated_output.csv")) if self.mlm_input.text() else "aggregated_output.csv",
                "CSV or Excel (*.csv *.xlsx)"
            )
            if file:
                self.output_input.setText(file)
        
        def load_metadata_dialog(self):
            file, _ = QFileDialog.getOpenFileName(self, "Select Metadata JSON", "", "JSON Files (*.json)")
            if file:
                self.load_metadata_from_file(Path(file))
        
        def load_metadata_from_file(self, json_path: Path):
            """Load settings from metadata JSON."""
            try:
                metadata = load_aggregation_metadata(json_path)
                tool = metadata.get("tool", "unknown")
                settings = metadata.get("settings", {})
                input_files = metadata.get("input_files", {})
                
                # Validate tool type
                supported_tools = ["RoBERTaMaskedLanguageModelVerbs", "MLMGroupAggregator"]
                if tool not in supported_tools:
                    QMessageBox.critical(
                        self,
                        "Unsupported Metadata Source",
                        f"This metadata is from '{tool}' which is not supported.\n\n"
                        f"MLMGroupAggregator accepts metadata from:\n"
                        f"  • RoBERTaMaskedLanguageModelVerbs\n"
                        f"  • MLMGroupAggregator\n\n"
                        f"Received: {tool}"
                    )
                    return
                if tool == "RoBERTaMaskedLanguageModelVerbs":
                    if metadata.get("output_file"):
                        self.mlm_input.setText(metadata.get("output_file", ""))
                    # User must provide group CSV and output file
                    self.topk_spin.setValue(settings.get("top_k", 0))
                    self.workers_spin.setValue(settings.get("workers", default_worker_count()))
                    self.lemma_col_input.setText(settings.get("lemma_col", "lemma"))
                    self.short_check.setChecked(settings.get("short", False))
                    self.count_check.setChecked(settings.get("include_count", False))
                # Handle MLMGroupAggregator metadata
                elif tool == "MLMGroupAggregator":
                    if input_files.get("mlm_csv"):
                        self.mlm_input.setText(input_files["mlm_csv"])
                    if input_files.get("group_csv"):
                        self.group_input.setText(input_files["group_csv"])
                    if metadata.get("output_file"):
                        self.output_input.setText(metadata.get("output_file", ""))
                    self.topk_spin.setValue(settings.get("top_k", 0))
                    self.workers_spin.setValue(settings.get("workers", default_worker_count()))
                    self.lemma_col_input.setText(settings.get("lemma_col", "lemma"))
                    self.short_check.setChecked(settings.get("short", False))
                    self.count_check.setChecked(settings.get("include_count", False))
                
                self.log(f"✓ Loaded settings from: {json_path}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to load metadata: {str(e)}")
        
        def log(self, message: str):
            """Add message to progress log."""
            self.progress_text.append(message)
            self.progress_text.verticalScrollBar().setValue(
                self.progress_text.verticalScrollBar().maximum()
            )

        def update_progress(self, processed: int, total: int):
            if total <= 0:
                # Keep progress bar in definite state to avoid animation
                self.progress_bar.setMaximum(1)
                self.progress_bar.setValue(0)
                self.progress_label.setText("Processing...")
                return
            progress_bytes = min(processed, total)
            permille = min(1000, int((progress_bytes / total) * 1000))
            percent = (progress_bytes / total) * 100
            self.progress_bar.setMaximum(1000)
            self.progress_bar.setValue(permille)
            self.progress_bar.setFormat(f"{percent:.1f}%")
            self.progress_label.setText(
                f"Processing: {_format_bytes(progress_bytes)} / {_format_bytes(total)} ({percent:.1f}%)"
            )
        
        def toggle_aggregation(self):
            """Start or stop the aggregation process."""
            # If worker is running, cancel it
            if self.worker and self.worker.isRunning():
                self.worker.cancel()
                return
            
            # Otherwise start new aggregation
            if not self.mlm_input.text():
                QMessageBox.warning(self, "Missing File", "Please select an MLM CSV file")
                return
            if not self.group_input.text():
                QMessageBox.warning(self, "Missing File", "Please select a Group CSV file")
                return
            if not self.output_input.text():
                QMessageBox.warning(self, "Missing File", "Please select an output file")
                return
            
            mlm_path = Path(self.mlm_input.text()).resolve()
            group_path = Path(self.group_input.text()).resolve()
            output_path = Path(self.output_input.text()).resolve()
            
            if not mlm_path.exists():
                QMessageBox.critical(self, "Error", f"MLM CSV not found: {mlm_path}")
                return
            if not group_path.exists():
                QMessageBox.critical(self, "Error", f"Group CSV not found: {group_path}")
                return
            
            self.progress_text.clear()
            self.log("=" * 60)
            self.log("MLM Group Aggregation")
            self.log("=" * 60)
            self.log(f"MLM CSV: {mlm_path}")
            self.log(f"Group CSV: {group_path}")
            self.log(f"Output: {output_path}")
            self.log(f"Worker Processes: {self.workers_spin.value()}")
            self.log("")
            
            self.start_btn.setEnabled(True)
            self.progress_bar.setMaximum(0)
            self.progress_bar.setValue(0)
            self.progress_bar.setFormat("")
            
            # Change button to "End Aggregation"
            self.start_btn.setText("End Aggregation")
            self.start_btn.setStyleSheet("background-color: #f44336; color: white; font-weight: bold; padding: 8px;")
            
            self.worker = AggregationWorker(
                mlm_path,
                group_path,
                output_path,
                self.topk_spin.value(),
                self.lemma_col_input.text(),
                self.short_check.isChecked(),
                self.count_check.isChecked(),
                "utf-8-sig",
                self.workers_spin.value(),
            )
            
            self.worker.progress_update.connect(self.log)
            self.worker.progress_value.connect(self.update_progress)
            self.worker.finished.connect(self.on_finished)
            self.worker.start()
        
        def on_finished(self, success: bool, message: str):
            """Handle aggregation completion."""
            self.start_btn.setEnabled(True)
            self.start_btn.setText("Start Aggregation")
            self.start_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 8px;")
            self.progress_bar.setMaximum(1)
            self.progress_bar.setValue(1)
            self.progress_bar.setFormat("Done" if success else "Cancelled")
            self.log("")
            self.log("=" * 60)
            self.log(message)
            self.log("=" * 60)
            
            if success:
                QMessageBox.information(self, "Success", message)
            else:
                QMessageBox.critical(self, "Error", message)
    
    app = QApplication(sys.argv)
    window = MLMGroupAggregatorGUI()
    window.show()
    sys.exit(app.exec())


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

def main() -> None:
    """Main entry point - choose GUI or CLI mode."""
    if len(sys.argv) == 1:
        print("Launching GUI mode...")
        print("(Use command-line arguments to run in CLI mode. Use --help for usage.)")
        run_gui()
    else:
        run_cli()


if __name__ == "__main__":
    main()
