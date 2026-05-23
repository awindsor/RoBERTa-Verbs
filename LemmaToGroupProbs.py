#!/usr/bin/env python3
"""
Aggregate group probabilities by lemma (auto-detect group columns).

Input: CSV produced by roberta_mlm_on_verbs.py with group columns appended.

Output:
  - If output filename ends with .csv  → write CSV
  - If output filename ends with .xlsx → write Excel (two sheets)

Excel sheets:
  1) lemma_to_groups:
        lemma → mean group probabilities (formatted as %) + max_prob + Group + Importance + count
        - bold the highest group percentage in each row
        - italicize lemma cells when the lemma is explicitly listed in at least
          one source group in the metadata, if available
        - color the lemma cell BLUE if the 2nd-highest group percentage is at least
          --second-threshold (default 0.50) times the highest
        - Group shows the max-probability group, or for blue lemmas a comma-separated
          list of all groups with probability >= second-threshold * max_prob
  2) groups_ranked:
        for each group: (lemma, pct) sorted by decreasing pct,
        with the lemma bolded in the group where it has the highest probability
        - italicize lemma cells when the lemma is explicitly listed in that
          group in the source group file metadata, if available
  3) group_overlap:
        square matrix of groups × groups where each cell is
        sum_lemma min(P(lemma in group_i), P(lemma in group_j))
  4) auto_groups:
        optional Excel-only sheet listing newly constructed groups in columns,
        using an importance cutoff and optional ambiguous-lemma exclusion

Auto-detection:
  - If --group-cols is provided, use it.
  - Otherwise infer group columns by taking all columns after the last prob_k column.

Requirements for Excel output:
  pip install openpyxl
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import logging
import os
import re
import sys
import time
import math
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Import openpyxl at module level to avoid threading issues
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

PROB_COL_RE = re.compile(r"^prob_(\d+)$")
OVERLAP_MEASURES = {"intersection", "correlation", "cosine"}
logger = logging.getLogger(__name__)
logger.addHandler(logging.NullHandler())


# ------------------------- metadata helpers -------------------------

def get_lemma_to_group_version_info() -> Dict[str, str]:
    """Return version info for this tool's dependencies."""
    try:
        import openpyxl  # type: ignore

        openpyxl_version = openpyxl.__version__
    except Exception:
        openpyxl_version = "not-installed"
    return {
        "python": f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}",
        "openpyxl": openpyxl_version,
    }


def compute_file_md5(path: Path) -> str:
    """Compute MD5 checksum of a file."""
    md5 = hashlib.md5()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            md5.update(chunk)
    return md5.hexdigest()


def reconstruct_command(input_csv: str, output_file: str, args: argparse.Namespace) -> str:
    """Reconstruct CLI command excluding defaults to aid reproducibility."""
    cmd = ["python", "LemmaToGroupProbs.py", input_csv, output_file]

    if args.lemma_col != "lemma":
        cmd.extend(["--lemma-col", args.lemma_col])
    if args.group_cols:
        cmd.extend(["--group-cols", *args.group_cols])
    if args.second_threshold != 0.50:
        cmd.extend(["--second-threshold", str(args.second_threshold)])
    if getattr(args, "auto_groups", False):
        cmd.append("--auto-groups")
        cmd.extend(["--importance-cutoff", str(args.importance_cutoff)])
    if not getattr(args, "ignore_ambiguous_auto_groups", True):
        cmd.append("--include-ambiguous-auto-groups")
    if getattr(args, "overlap_measure", "intersection") != "intersection":
        cmd.extend(["--overlap-measure", args.overlap_measure])
    if getattr(args, "vba_links", False):
        cmd.append("--vba-links")
    if args.encoding != "utf-8":
        cmd.extend(["--encoding", args.encoding])
    if args.load_metadata:
        cmd.extend(["--load-metadata", args.load_metadata])

    return " ".join(cmd)


def save_metadata(
    output_path: Path,
    input_path: Path,
    input_checksum: str,
    output_checksum: str,
    args: Dict[str, Any],
    stats: Dict[str, Any],
    group_cols: List[str],
    source_metadata: Optional[Dict[str, Any]] = None,
    command: Optional[str] = None,
) -> None:
    """Write metadata JSON next to the output file."""
    json_path = output_path.with_suffix(".json")
    metadata: Dict[str, Any] = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "tool": "LemmaToGroupProbs",
        "versions": get_lemma_to_group_version_info(),
        "input_file": str(input_path),
        "input_checksum": input_checksum,
        "output_file": str(output_path),
        "output_checksum": output_checksum,
        "command": command,
        "settings": {
            "lemma_col": args.get("lemma_col", "lemma"),
            "group_cols": group_cols,
            "second_threshold": args.get("second_threshold", 0.50),
            "auto_groups": args.get("auto_groups", False),
            "importance_cutoff": args.get("importance_cutoff", 0.0),
            "ignore_ambiguous_auto_groups": args.get("ignore_ambiguous_auto_groups", True),
            "overlap_measure": args.get("overlap_measure", "intersection"),
            "encoding": args.get("encoding", "utf-8"),
            "output_format": output_path.suffix.lower().lstrip("."),
            "vba_links": args.get("vba_links", False),
        },
        "statistics": stats,
    }

    # Handle source metadata chaining:
    # - If loading from MLMGroupAggregator: include the entire JSON as source_metadata
    # - If loading from LemmaToGroupProbs: include only its source_metadata (preserve chain, avoid nesting)
    if source_metadata:
        source_tool = source_metadata.get("tool")
        if source_tool == "MLMGroupAggregator":
            # Include the entire MLMGroupAggregator JSON
            metadata["source_metadata"] = source_metadata
        elif source_tool == "LemmaToGroupProbs":
            # Include only the source_metadata from the previous LemmaToGroupProbs run
            # This preserves the chain without creating nested LemmaToGroupProbs entries
            if "source_metadata" in source_metadata:
                metadata["source_metadata"] = source_metadata["source_metadata"]

    with json_path.open("w", encoding="utf-8") as f:
        json.dump(metadata, f, indent=2)


# ------------------------- helpers -------------------------

def infer_group_cols(fieldnames: List[str]) -> List[str]:
    """Infer group columns as those appearing after the last prob_k column."""
    last_prob_idx = -1
    for i, name in enumerate(fieldnames):
        if PROB_COL_RE.match(name.strip()):
            last_prob_idx = i
    if last_prob_idx == -1:
        return []
    return [clean_group_col_name(c) for c in fieldnames[last_prob_idx + 1:] if c.strip()]


def clean_group_col_name(name: str) -> str:
    """Normalize a group column name for display/metadata matching."""
    stripped = name.strip()
    if stripped.startswith("group_prob_"):
        return stripped[len("group_prob_"):]
    return stripped


def safe_sheet_title(s: str) -> str:
    bad = r'[]:*?/\\'
    out = "".join("_" if ch in bad else ch for ch in s)
    return out[:31]


def extract_group_membership_from_metadata(
    metadata: Optional[Dict[str, Any]],
) -> Dict[str, set[str]]:
    """Extract group -> lemma membership mapping from chained metadata, if present."""
    if not metadata:
        return {}

    groups = metadata.get("groups")
    if not isinstance(groups, dict):
        nested = metadata.get("source_metadata")
        if isinstance(nested, dict):
            return extract_group_membership_from_metadata(nested)
        return {}

    membership: Dict[str, set[str]] = {}
    for group_name, lemmas in groups.items():
        if isinstance(lemmas, list):
            membership[str(group_name)] = {
                str(lemma).strip() for lemma in lemmas if str(lemma).strip()
            }
    return membership


def load_input_sidecar_metadata(input_path: Path) -> Optional[Dict[str, Any]]:
    """Load metadata JSON adjacent to the current input CSV, if present."""
    meta_path = input_path.with_suffix(".json")
    if not meta_path.exists():
        return None
    try:
        with meta_path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def compute_group_matrix(
    group_cols: List[str],
    counts: Dict[str, int],
    means: Dict[str, Dict[str, float]],
    overlap_measure: str,
) -> Dict[str, Dict[str, float]]:
    """Compute a group-by-group matrix using the requested similarity measure."""
    if overlap_measure not in OVERLAP_MEASURES:
        raise ValueError(f"Unsupported overlap measure: {overlap_measure}")

    lemmas = list(counts.keys())
    matrix: Dict[str, Dict[str, float]] = {}

    for g1 in group_cols:
        row: Dict[str, float] = {}
        v1 = [means[g1][lemma] for lemma in lemmas]
        mean1 = sum(v1) / len(v1) if v1 else 0.0
        norm1 = math.sqrt(sum(x * x for x in v1))

        for g2 in group_cols:
            v2 = [means[g2][lemma] for lemma in lemmas]

            if overlap_measure == "intersection":
                row[g2] = sum(min(a, b) for a, b in zip(v1, v2))
            elif overlap_measure == "correlation":
                mean2 = sum(v2) / len(v2) if v2 else 0.0
                num = sum((a - mean1) * (b - mean2) for a, b in zip(v1, v2))
                den1 = math.sqrt(sum((a - mean1) ** 2 for a in v1))
                den2 = math.sqrt(sum((b - mean2) ** 2 for b in v2))
                row[g2] = num / (den1 * den2) if den1 > 0 and den2 > 0 else 0.0
            else:  # cosine
                norm2 = math.sqrt(sum(x * x for x in v2))
                dot = sum(a * b for a, b in zip(v1, v2))
                row[g2] = dot / (norm1 * norm2) if norm1 > 0 and norm2 > 0 else 0.0

        matrix[g1] = row

    return matrix


def format_group_label(
    lemma: str,
    group_cols: List[str],
    means: Dict[str, Dict[str, float]],
    second_threshold: float,
) -> Tuple[str, str, float, float]:
    """Return group label plus max-group stats for a lemma."""
    group_vals = [(g, means[g][lemma]) for g in group_cols]
    group_vals.sort(key=lambda x: x[1], reverse=True)
    max_g, max_v = group_vals[0]
    second_v = group_vals[1][1] if len(group_vals) > 1 else 0.0

    if max_v > 0 and second_v >= (second_threshold * max_v):
        group_label = ", ".join(
            g for g, v in group_vals if v >= (second_threshold * max_v)
        )
    else:
        group_label = max_g

    return group_label, max_g, max_v, second_v


def excel_string_literal(value: str) -> str:
    """Return an Excel string literal with embedded quotes escaped."""
    return '"' + value.replace('"', '""') + '"'


def build_internal_hyperlink_formula(display_text: str, target_sheet_title: str) -> str:
    """Build a formula hyperlink that finds the lemma row dynamically on Sheet 1."""
    display_literal = excel_string_literal(display_text)
    sheet_literal = target_sheet_title.replace("'", "''")
    return (
        f'=HYPERLINK('
        f'"#\'{sheet_literal}\'!A"&MATCH({display_literal},\'{sheet_literal}\'!A:A,0),'
        f'{display_literal}'
        f')'
    )


def output_mode_from_suffix(suffix: str) -> str:
    """Map an output suffix to the GUI output mode."""
    suffix = suffix.lower()
    if suffix == ".csv":
        return ".csv"
    if suffix == ".tsv":
        return ".tsv"
    return "excel"


def compute_importance(count: int, max_prob: float) -> float:
    """Compute lemma importance from count and max probability."""
    return math.log10(count) * max_prob


def build_auto_group_assignments(
    group_cols: List[str],
    counts: Dict[str, int],
    means: Dict[str, Dict[str, float]],
    second_threshold: float,
    importance_cutoff: float,
    ignore_ambiguous: bool,
) -> Dict[str, List[str]]:
    """Assign lemmas to auto-groups based on importance and ambiguity rules."""
    assignments: Dict[str, List[Tuple[str, float, float]]] = {g: [] for g in group_cols}

    for lemma in sorted(counts):
        _group_label, max_g, max_v, second_v = format_group_label(
            lemma, group_cols, means, second_threshold
        )
        importance = compute_importance(counts[lemma], max_v)
        is_ambiguous = max_v > 0 and second_v >= (second_threshold * max_v)

        if importance < importance_cutoff:
            continue
        if is_ambiguous and ignore_ambiguous:
            continue

        assignments[max_g].append((lemma, importance, max_v))

    ordered: Dict[str, List[str]] = {}
    for group_name in group_cols:
        items = assignments[group_name]
        items.sort(key=lambda x: (-x[1], -x[2], x[0]))
        ordered[group_name] = [lemma for lemma, _importance, _max_v in items]

    return ordered


def append_auto_groups_sheet(
    wb: Workbook,
    group_cols: List[str],
    counts: Dict[str, int],
    means: Dict[str, Dict[str, float]],
    second_threshold: float,
    importance_cutoff: float,
    ignore_ambiguous: bool,
    header_font: Optional[Font] = None,
    header_align: Optional[Alignment] = None,
) -> None:
    """Append the optional auto_groups sheet."""
    ws = wb.create_sheet(title=safe_sheet_title("auto_groups"))
    assignments = build_auto_group_assignments(
        group_cols=group_cols,
        counts=counts,
        means=means,
        second_threshold=second_threshold,
        importance_cutoff=importance_cutoff,
        ignore_ambiguous=ignore_ambiguous,
    )

    ws.append(group_cols)
    if header_font is not None and header_align is not None:
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_align

    max_len = max((len(assignments[g]) for g in group_cols), default=0)
    for row_idx in range(max_len):
        row = []
        for group_name in group_cols:
            if row_idx < len(assignments[group_name]):
                row.append(assignments[group_name][row_idx])
            else:
                row.append(None)
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


# ------------------------- Excel writer -------------------------

def write_excel(
    xlsx_path: str,
    lemma_col: str,
    group_cols: List[str],
    counts: Dict[str, int],
    means: Dict[str, Dict[str, float]],
    best_group_for_lemma: Dict[str, str],
    second_threshold: float,
    overlap_measure: str,
    group_membership: Optional[Dict[str, set[str]]] = None,
    use_vba_links: bool = False,
    auto_groups: bool = False,
    importance_cutoff: float = 0.0,
    ignore_ambiguous_auto_groups: bool = True,
) -> None:
    """
    Write Excel output. For large datasets, use CSV instead - Excel formatting
    is memory-intensive for hundreds of thousands of rows.
    """
    logger.debug("write_excel: entered")
    logger.debug("write_excel: xlsx_path=%s", xlsx_path)
    logger.debug("write_excel: total lemmas=%d, groups=%d", len(counts), len(group_cols))
    
    if not OPENPYXL_AVAILABLE:
        logger.error("write_excel: openpyxl not available")
        raise SystemExit(
            "Excel output requested but openpyxl is not installed. Install with: pip install openpyxl"
        )
    
    logger.debug("write_excel: openpyxl is available (imported at module level)")

    overlap_matrix = compute_group_matrix(group_cols, counts, means, overlap_measure)
    if use_vba_links:
        logger.warning(
            "write_excel: VBA link mode requested, but workbook macros are not generated in-code; "
            "writing .xlsm-compatible workbook with formula-based internal links."
        )

    # For very large datasets, warn and limit formatting
    total_lemmas = len(counts)
    
    try:
        if total_lemmas > 100000:
            logger.warning(
                "write_excel: %d lemmas is very large for Excel. Consider using CSV output instead.",
                total_lemmas,
            )
            logger.info("write_excel: writing Excel without formatting to reduce memory usage")
            
            # Simple write without formatting for large datasets
            logger.debug("write_excel: creating workbook")
            wb = Workbook()
            ws1 = wb.active
            ws1.title = safe_sheet_title("lemma_to_groups")
            
            # Header
            logger.debug("write_excel: writing Sheet 1 header")
            header = [lemma_col] + group_cols + ["max_prob", "Group", "Importance", "count"]
            ws1.append(header)
            
            # Data rows only
            logger.debug("write_excel: writing %d rows to Sheet 1", total_lemmas)
            for i, lemma in enumerate(sorted(counts)):
                if i % 10000 == 0 and i > 0:
                    logger.info("write_excel: written %d/%d rows", i, total_lemmas)
                group_label, _max_g, max_prob, _second_v = format_group_label(
                    lemma, group_cols, means, second_threshold
                )
                importance = compute_importance(counts[lemma], max_prob)
                row_vals = [lemma] + [means[g][lemma] for g in group_cols] + [max_prob, group_label, importance, counts[lemma]]
                ws1.append(row_vals)
            
            # Sheet 2
            logger.debug("write_excel: creating Sheet 2")
            ws2 = wb.create_sheet(title=safe_sheet_title("groups_ranked"))
            
            logger.debug("write_excel: sorting groups for Sheet 2")
            per_group_sorted_simple: Dict[str, List[Tuple[str, float]]] = {}
            max_len = 0
            for g in group_cols:
                items = [(lemma, means[g][lemma]) for lemma in counts]
                items.sort(key=lambda x: x[1], reverse=True)
                per_group_sorted_simple[g] = items
                max_len = max(max_len, len(items))
            
            logger.debug("write_excel: writing Sheet 2 header")
            header2_simple: List[str] = []
            for g in group_cols:
                header2_simple.extend([f"{g}_lemma", f"{g}_pct"])
            ws2.append(header2_simple)
            
            logger.debug("write_excel: writing %d rows to Sheet 2", max_len)
            sheet1_title = ws1.title
            for r in range(max_len):
                if r % 10000 == 0 and r > 0:
                    logger.info("write_excel: written %d/%d rows", r, max_len)
                row_data = []
                for g in group_cols:
                    if r < len(per_group_sorted_simple[g]):
                        lemma, pct = per_group_sorted_simple[g][r]
                        row_data.extend([build_internal_hyperlink_formula(lemma, sheet1_title), pct])
                    else:
                        row_data.extend([None, None])
                ws2.append(row_data)

            # Sheet 3
            logger.debug("write_excel: creating Sheet 3")
            ws3 = wb.create_sheet(title=safe_sheet_title("group_overlap"))
            ws3.append(["group"] + group_cols)
            for g1 in group_cols:
                ws3.append([g1] + [overlap_matrix[g1][g2] for g2 in group_cols])

            if auto_groups:
                append_auto_groups_sheet(
                    wb=wb,
                    group_cols=group_cols,
                    counts=counts,
                    means=means,
                    second_threshold=second_threshold,
                    importance_cutoff=importance_cutoff,
                    ignore_ambiguous=ignore_ambiguous_auto_groups,
                )
            
            logger.debug("write_excel: saving workbook to %s", xlsx_path)
            wb.save(xlsx_path)
            logger.info("write_excel: Excel file saved successfully")
            return

        logger.debug("write_excel: using formatted Excel output")
        wb = Workbook()
        bold_font = Font(bold=True)
        italic_font = Font(italic=True)
        blue_font = Font(color="0000FF")  # blue text
        italic_blue_font = Font(italic=True, color="0000FF")
        bold_italic_font = Font(bold=True, italic=True)
        bold_blue_font = Font(bold=True, color="0000FF")
        bold_italic_blue_font = Font(bold=True, italic=True, color="0000FF")
        header_font = Font(bold=True)
        header_align = Alignment(vertical="center")

        # ---------- Sheet 1: lemma_to_groups ----------
        ws1 = wb.active
        ws1.title = safe_sheet_title("lemma_to_groups")

        header = [lemma_col] + group_cols + ["max_prob", "Group", "Importance", "count"]
        ws1.append(header)
        for cell in ws1[1]:
            cell.font = header_font
            cell.alignment = header_align

        # Write rows + apply formatting rules
        # Columns:
        #   lemma = 1
        #   groups = 2..(1+len(group_cols))
        #   count = last
        ambiguous_lemmas = set()
        for lemma in sorted(counts):
            group_label, max_g, max_v, second_v = format_group_label(
                lemma, group_cols, means, second_threshold
            )
            max_prob = max_v
            importance = compute_importance(counts[lemma], max_prob)
            row_vals = [lemma] + [means[g][lemma] for g in group_cols] + [max_prob, group_label, importance, counts[lemma]]
            ws1.append(row_vals)
            r = ws1.max_row
            is_in_any_source_group = (
                group_membership is not None and
                any(lemma in members for members in group_membership.values())
            )

            # Bold the max group cell in this row
            max_idx = group_cols.index(max_g)  # 0-based within group_cols
            max_col = 2 + max_idx  # Excel column index for the group cell
            ws1.cell(row=r, column=max_col).font = bold_font

            lemma_cell = ws1.cell(row=r, column=1)
            if is_in_any_source_group:
                lemma_cell.font = italic_font

            # If 2nd-highest >= threshold * highest, color lemma cell blue
            # Avoid dividing by zero: if max_v == 0, treat as "not ambiguous"
            if max_v > 0 and second_v >= (second_threshold * max_v):
                lemma_cell.font = italic_blue_font if is_in_any_source_group else blue_font
                ambiguous_lemmas.add(lemma)

        # Percent formatting for group columns + max_prob
        for j in range(2, 3 + len(group_cols)):
            for i in range(2, ws1.max_row + 1):
                ws1.cell(row=i, column=j).number_format = "0.00%"

        ws1.freeze_panes = "A2"
        ws1.auto_filter.ref = f"A1:{get_column_letter(ws1.max_column)}{ws1.max_row}"

        # ---------- Sheet 2: groups_ranked ----------
        ws2 = wb.create_sheet(title=safe_sheet_title("groups_ranked"))

        per_group_sorted: Dict[str, List[Tuple[str, float]]] = {}
        max_len = 0
        for g in group_cols:
            items = [(lemma, means[g][lemma]) for lemma in counts]
            items.sort(key=lambda x: x[1], reverse=True)
            per_group_sorted[g] = items
            max_len = max(max_len, len(items))

        header2: List[str] = []
        for g in group_cols:
            header2.extend([f"{g}_lemma", f"{g}_pct"])
        ws2.append(header2)
        for cell in ws2[1]:
            cell.font = header_font
            cell.alignment = header_align

        sheet1_title = ws1.title
        for r in range(max_len):
            row_data = []
            for k, g in enumerate(group_cols):
                if r < len(per_group_sorted[g]):
                    lemma, pct = per_group_sorted[g][r]
                    row_data.extend([build_internal_hyperlink_formula(lemma, sheet1_title), pct])
                else:
                    row_data.extend([None, None])
            ws2.append(row_data)
            excel_row = ws2.max_row

            for k, g in enumerate(group_cols):
                if r < len(per_group_sorted[g]):
                    lemma, pct = per_group_sorted[g][r]
                    lemma_col_idx = 1 + 2 * k
                    
                    cell_lemma = ws2.cell(row=excel_row, column=lemma_col_idx)
                    is_best = best_group_for_lemma.get(lemma) == g
                    is_ambiguous = lemma in ambiguous_lemmas
                    is_in_source_group = (
                        group_membership is not None and
                        lemma in group_membership.get(g, set())
                    )
                    if is_best and is_ambiguous and is_in_source_group:
                        cell_lemma.font = bold_italic_blue_font
                    elif is_best and is_ambiguous:
                        cell_lemma.font = bold_blue_font
                    elif is_best and is_in_source_group:
                        cell_lemma.font = bold_italic_font
                    elif is_ambiguous and is_in_source_group:
                        cell_lemma.font = italic_blue_font
                    elif is_best:
                        cell_lemma.font = bold_font
                    elif is_in_source_group:
                        cell_lemma.font = italic_font
                    elif is_ambiguous:
                        cell_lemma.font = blue_font

        for col in range(2, 2 * len(group_cols) + 1, 2):
            for i in range(2, ws2.max_row + 1):
                ws2.cell(row=i, column=col).number_format = "0.00%"

        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = f"A1:{get_column_letter(ws2.max_column)}{ws2.max_row}"

        # ---------- Sheet 3: group_overlap ----------
        ws3 = wb.create_sheet(title=safe_sheet_title("group_overlap"))
        ws3.append(["group"] + group_cols)
        for cell in ws3[1]:
            cell.font = header_font
            cell.alignment = header_align

        for g1 in group_cols:
            ws3.append([g1] + [overlap_matrix[g1][g2] for g2 in group_cols])
            r = ws3.max_row
            ws3.cell(row=r, column=1).font = header_font
            ws3.cell(row=r, column=1).alignment = header_align

        matrix_number_format = "0.00%" if overlap_measure == "intersection" else "0.0000"
        for col in range(2, ws3.max_column + 1):
            for i in range(2, ws3.max_row + 1):
                ws3.cell(row=i, column=col).number_format = matrix_number_format

        ws3.freeze_panes = "B2"
        ws3.auto_filter.ref = f"A1:{get_column_letter(ws3.max_column)}{ws3.max_row}"

        if auto_groups:
            append_auto_groups_sheet(
                wb=wb,
                group_cols=group_cols,
                counts=counts,
                means=means,
                second_threshold=second_threshold,
                importance_cutoff=importance_cutoff,
                ignore_ambiguous=ignore_ambiguous_auto_groups,
                header_font=header_font,
                header_align=header_align,
            )

        logger.debug("write_excel: saving workbook (formatted version)")
        wb.save(xlsx_path)
        logger.info("write_excel: Excel file saved successfully")
    
    except Exception as e:
        logger.exception("write_excel: %s: %s", type(e).__name__, e)
        raise


# ------------------------- main -------------------------

def run_cli(
    argv: Optional[List[str]] = None,
    on_progress=None,
    on_progress_value=None,
    total_rows_expected: Optional[int] = None,
) -> None:
    """
    Run in CLI mode.
    
    Args:
        argv: Command-line arguments (if None, uses sys.argv)
        on_progress: Optional callback(msg) for progress messages
    """
    def log(msg: str):
        logger.info("%s", msg)
        if on_progress:
            on_progress(msg)
        else:
            print(msg)
    
    ap = argparse.ArgumentParser(description="Aggregate group probabilities by lemma")
    ap.add_argument("input_csv", help="Input CSV file with group probability columns")
    ap.add_argument("output", help="Output filename (.csv, .tsv, .xlsx, or .xlsm)")
    ap.add_argument("--lemma-col", default="lemma", help="Lemma column name (default: lemma)")
    ap.add_argument(
        "--group-cols",
        nargs="+",
        default=None,
        help="Optional explicit group column names; otherwise inferred from header",
    )
    ap.add_argument(
        "--second-threshold",
        type=float,
        default=0.50,
        help="Color lemma blue if 2nd-best >= threshold * best (default: 0.50)",
    )
    ap.add_argument(
        "--auto-groups",
        action="store_true",
        help="Add an Excel-only auto_groups sheet using importance-based assignment.",
    )
    ap.add_argument(
        "--importance-cutoff",
        type=float,
        default=0.0,
        help="Minimum importance for auto-group assignment (default: 0.0).",
    )
    ap.add_argument(
        "--include-ambiguous-auto-groups",
        action="store_true",
        help="Include ambiguous lemmas in auto_groups by assigning them to their top group.",
    )
    ap.add_argument(
        "--overlap-measure",
        default="intersection",
        choices=sorted(OVERLAP_MEASURES),
        help="Matrix measure for the third sheet: intersection, correlation, or cosine (default: intersection)",
    )
    ap.add_argument(
        "--vba-links",
        action="store_true",
        help="Write Excel output as .xlsm and reserve VBA link mode; currently still uses formula-based hyperlinks.",
    )
    ap.add_argument("--encoding", default="utf-8-sig", help="Input file encoding (default: utf-8-sig to handle BOM)")
    ap.add_argument("--load-metadata", help="Load settings from metadata JSON (e.g., MLMGroupAggregator or previous run)")
    parsed = ap.parse_args(args=argv)
    explicit_group_cols = parsed.group_cols is not None

    if not (0.0 <= parsed.second_threshold <= 1.0):
        raise SystemExit("--second-threshold must be between 0 and 1.")
    if parsed.importance_cutoff < 0.0:
        raise SystemExit("--importance-cutoff must be non-negative.")

    if parsed.vba_links and parsed.output.lower().endswith(".xlsx"):
        parsed.output = str(Path(parsed.output).with_suffix(".xlsm"))

    out_ext = os.path.splitext(parsed.output)[1].lower()
    if out_ext not in {".csv", ".tsv", ".xlsx", ".xlsm"}:
        raise SystemExit("Output filename must end with .csv, .tsv, .xlsx, or .xlsm")

    source_metadata: Optional[Dict[str, Any]] = None
    if parsed.load_metadata:
        meta_path = Path(parsed.load_metadata)
        if not meta_path.exists():
            raise SystemExit(f"Metadata file not found: {meta_path}")
        with meta_path.open("r", encoding="utf-8") as f:
            source_metadata = json.load(f)
        # Try to pull defaults from metadata if not explicitly provided
        if source_metadata is not None:
            if parsed.group_cols is None:
                if "groups" in source_metadata and isinstance(source_metadata["groups"], dict):
                    parsed.group_cols = list(source_metadata["groups"].keys())
            if not parsed.lemma_col and "settings" in source_metadata:
                parsed.lemma_col = source_metadata.get("settings", {}).get("lemma_col", parsed.lemma_col)
            if (
                parsed.overlap_measure == "intersection" and
                "settings" in source_metadata and
                isinstance(source_metadata.get("settings"), dict)
            ):
                parsed.overlap_measure = source_metadata["settings"].get("overlap_measure", parsed.overlap_measure)
            settings = source_metadata.get("settings", {})
            if isinstance(settings, dict):
                if not parsed.auto_groups:
                    parsed.auto_groups = bool(settings.get("auto_groups", parsed.auto_groups))
                if parsed.importance_cutoff == 0.0:
                    parsed.importance_cutoff = float(settings.get("importance_cutoff", parsed.importance_cutoff))
                if not parsed.include_ambiguous_auto_groups:
                    parsed.include_ambiguous_auto_groups = not bool(
                        settings.get("ignore_ambiguous_auto_groups", True)
                    )

    input_path = Path(parsed.input_csv)
    output_path = Path(parsed.output)
    input_sidecar_metadata = load_input_sidecar_metadata(input_path)
    group_metadata_source = source_metadata
    if (
        input_sidecar_metadata is not None and
        input_sidecar_metadata.get("tool") == "MLMGroupAggregator" and
        isinstance(input_sidecar_metadata.get("groups"), dict)
    ):
        group_metadata_source = input_sidecar_metadata
        if not explicit_group_cols:
            parsed.group_cols = list(input_sidecar_metadata["groups"].keys())

    group_membership = extract_group_membership_from_metadata(group_metadata_source)

    start_time = time.time()
    total_rows = 0

    with input_path.open(newline="", encoding=parsed.encoding) as fin:
        reader = csv.DictReader(fin)
        if reader.fieldnames is None:
            raise SystemExit("Input CSV has no header.")

        if parsed.lemma_col not in reader.fieldnames:
            raise SystemExit(f"Missing lemma column {parsed.lemma_col!r}")

        group_cols = parsed.group_cols or infer_group_cols(list(reader.fieldnames))
        if not group_cols:
            raise SystemExit("Could not infer group columns. Provide --group-cols explicitly.")
        
        # Build mapping from clean group names to actual column names (which may have BOM)
        group_col_mapping = {}
        for gc in group_cols:
            matching = [col for col in reader.fieldnames if clean_group_col_name(col) == gc]
            if matching:
                # Use the first match (should be exactly one)
                group_col_mapping[gc] = matching[0]
            else:
                raise SystemExit(f"Group column '{gc}' not found in CSV fieldnames")
        
        # Use the actual column names from the CSV for lookups
        actual_group_cols = [group_col_mapping[gc] for gc in group_cols]
        
        sums: Dict[str, Dict[str, float]] = {g: defaultdict(float) for g in group_cols}
        counts: Dict[str, int] = defaultdict(int)

        for row in reader:
            total_rows += 1
            
            # Report progress to GUI if requested (no row-count log spam)
            if on_progress_value and total_rows % 1000 == 0:
                on_progress_value(total_rows, total_rows_expected)
            
            lemma = (row.get(parsed.lemma_col) or "").strip()
            if not lemma:
                continue

            counts[lemma] += 1
            
            for clean_g, actual_g in zip(group_cols, actual_group_cols):
                val = (row.get(actual_g) or "").strip()
                if not val:
                    continue
                try:
                    sums[clean_g][lemma] += float(val)
                except ValueError:
                    pass

    means: Dict[str, Dict[str, float]] = {
        g: {lemma: (sums[g][lemma] / counts[lemma]) for lemma in counts}
        for g in group_cols
    }

    log(f"Computed means for {len(counts)} unique lemmas")

    best_group_for_lemma: Dict[str, str] = {}
    log("Computing best group for each lemma...")
    for lemma in counts:
        best_group_for_lemma[lemma] = max(group_cols, key=lambda g: means[g][lemma])

    log(f"Writing output to {output_path}...")
    log(f"Output format: {out_ext}")
    
    if out_ext in {".csv", ".tsv"}:
        delimiter = "," if out_ext == ".csv" else "\t"
        log(f"Starting {'CSV' if out_ext == '.csv' else 'TSV'} write...")
        with output_path.open("w", newline="", encoding=parsed.encoding) as fout:
            log("Delimited text file opened for writing")
            writer = csv.DictWriter(
                fout,
                fieldnames=[parsed.lemma_col] + group_cols + ["max_prob", "Group", "Importance", "count"],
                delimiter=delimiter,
            )
            log("Writing header...")
            writer.writeheader()
            log(f"Writing {len(counts)} data rows...")
            for i, lemma in enumerate(sorted(counts)):
                if i % 1000 == 0 and i > 0:
                    log(f"  Written {i}/{len(counts)} rows...")
                group_label, _max_g, max_prob, _second_v = format_group_label(
                    lemma, group_cols, means, parsed.second_threshold
                )
                row = {
                    parsed.lemma_col: lemma,
                    "max_prob": f"{max_prob:.10g}",
                    "Group": group_label,
                    "Importance": f"{compute_importance(counts[lemma], max_prob):.10g}",
                    "count": counts[lemma],
                }
                for g in group_cols:
                    row[g] = f"{means[g][lemma]:.10g}"
                writer.writerow(row)
        log(f"✓ Wrote {output_path}")
    else:
        log(f"Calling write_excel with {len(counts)} lemmas...")
        try:
            write_excel(
                xlsx_path=str(output_path),
                lemma_col=parsed.lemma_col,
                group_cols=group_cols,
                counts=counts,
                means=means,
                best_group_for_lemma=best_group_for_lemma,
                second_threshold=parsed.second_threshold,
                overlap_measure=parsed.overlap_measure,
                group_membership=group_membership,
                use_vba_links=parsed.vba_links,
                auto_groups=parsed.auto_groups,
                importance_cutoff=parsed.importance_cutoff,
                ignore_ambiguous_auto_groups=not parsed.include_ambiguous_auto_groups,
            )
            log(f"✓ Wrote {output_path}")
        except Exception as e:
            log(f"ERROR in write_excel: {type(e).__name__}: {e}")
            import traceback
            log(traceback.format_exc())
            raise

    log(f"Output written successfully. Computing checksums...")
    if on_progress_value:
        on_progress_value(total_rows, total_rows_expected)
    elapsed = time.time() - start_time

    # Metadata - compute checksums
    log(f"Computing MD5 checksum for input file: {input_path}")
    try:
        input_checksum = compute_file_md5(input_path)
        log(f"Input checksum: {input_checksum}")
    except Exception as e:
        log(f"Warning: Could not compute input checksum: {e}")
        input_checksum = "error"
    
    log(f"Computing MD5 checksum for output file: {output_path}")
    try:
        output_checksum = compute_file_md5(output_path)
        log(f"Output checksum: {output_checksum}")
    except Exception as e:
        log(f"Warning: Could not compute output checksum: {e}")
        output_checksum = "error"
    
    log("Creating statistics dictionary...")
    stats = {
        "lemmas": len(counts),
        "total_rows": total_rows,
        "elapsed_seconds": elapsed,
    }
    log(f"Reconstructing command...")
    command = reconstruct_command(parsed.input_csv, parsed.output, parsed)
    
    log("Saving metadata to JSON...")
    try:
        save_metadata(
            output_path=output_path,
            input_path=input_path,
            input_checksum=input_checksum,
            output_checksum=output_checksum,
            args={
                "lemma_col": parsed.lemma_col,
                "group_cols": group_cols,
                "second_threshold": parsed.second_threshold,
                "auto_groups": parsed.auto_groups,
                "importance_cutoff": parsed.importance_cutoff,
                "ignore_ambiguous_auto_groups": not parsed.include_ambiguous_auto_groups,
                "overlap_measure": parsed.overlap_measure,
                "vba_links": parsed.vba_links,
                "encoding": parsed.encoding,
            },
            stats=stats,
            group_cols=group_cols,
            source_metadata=source_metadata,
            command=command,
        )
        log(f"✓ Saved metadata to {output_path.with_suffix('.json')}")
    except Exception as e:
        log(f"✗ Error saving metadata: {e}")
        import traceback
        log(traceback.format_exc())
    


def run_gui() -> None:
    try:
        from PySide6.QtCore import QThread, Signal
        from PySide6.QtWidgets import (
            QApplication,
            QCheckBox,
            QComboBox,
            QFileDialog,
            QGridLayout,
            QGroupBox,
            QHBoxLayout,
            QLabel,
            QLineEdit,
            QMainWindow,
            QMessageBox,
            QProgressBar,
            QPushButton,
            QSpinBox,
            QTextEdit,
            QVBoxLayout,
            QWidget,
        )
    except ImportError:
        print("Error: PySide6 is required for GUI mode. Install with: pip install PySide6")
        sys.exit(1)

    class Worker(QThread):
        progress = Signal(str)
        progress_value = Signal(int, int)
        finished = Signal(bool, str)

        def __init__(self, input_path: Path, output_path: Path, lemma_col: str, group_cols: Optional[List[str]], second_threshold: float, overlap_measure: str, auto_groups: bool, importance_cutoff: float, ignore_ambiguous_auto_groups: bool, use_vba_links: bool, encoding: str, metadata_path: Optional[Path]):
            super().__init__()
            self.input_path = input_path
            self.output_path = output_path
            self.lemma_col = lemma_col
            self.group_cols = group_cols
            self.second_threshold = second_threshold
            self.overlap_measure = overlap_measure
            self.auto_groups = auto_groups
            self.importance_cutoff = importance_cutoff
            self.ignore_ambiguous_auto_groups = ignore_ambiguous_auto_groups
            self.use_vba_links = use_vba_links
            self.encoding = encoding
            self.metadata_path = metadata_path

        def run(self):
            try:
                self.progress.emit("Starting aggregation...")
                # Count rows for determinate progress
                total_rows = 0
                try:
                    with self.input_path.open(newline="", encoding=self.encoding) as fin:
                        reader = csv.reader(fin)
                        next(reader, None)  # header
                        for _ in reader:
                            total_rows += 1
                    self.progress_value.emit(0, total_rows)
                except Exception:
                    total_rows = 0
                    self.progress_value.emit(0, 0)
                argv = [
                    str(self.input_path),
                    str(self.output_path),
                    "--lemma-col",
                    self.lemma_col,
                    "--second-threshold",
                    str(self.second_threshold),
                    "--overlap-measure",
                    self.overlap_measure,
                    "--encoding",
                    self.encoding,
                ]
                if self.auto_groups:
                    argv.extend(["--auto-groups", "--importance-cutoff", str(self.importance_cutoff)])
                    if not self.ignore_ambiguous_auto_groups:
                        argv.append("--include-ambiguous-auto-groups")
                if self.use_vba_links:
                    argv.append("--vba-links")
                if self.group_cols:
                    argv.extend(["--group-cols", *self.group_cols])
                if self.metadata_path:
                    argv.extend(["--load-metadata", str(self.metadata_path)])
                
                self.progress.emit(f"Input: {argv[0]}")
                self.progress.emit(f"Output: {argv[1]}")
                
                def progress_callback(msg):
                    """Handle progress messages."""
                    if msg is not None:
                        self.progress.emit(msg)
                
                def progress_value_callback(current, total):
                    self.progress_value.emit(current, total if total is not None else total_rows)

                run_cli(
                    argv,
                    on_progress=progress_callback,
                    on_progress_value=progress_value_callback,
                    total_rows_expected=total_rows if total_rows > 0 else None,
                )
                self.progress.emit("✓ Aggregation complete")
                self.finished.emit(True, "Completed successfully")
                
            except SystemExit as e:
                msg = f"Exit code {e.code}: Check the log for details"
                self.progress.emit(f"✗ {msg}")
                self.finished.emit(False, msg)
            except Exception as e:
                import traceback
                msg = f"Error: {str(e)}"
                self.progress.emit(f"✗ {msg}")
                self.progress.emit(traceback.format_exc())
                self.finished.emit(False, msg)

    class MainWindow(QMainWindow):
        def __init__(self):
            super().__init__()
            self.setWindowTitle("Lemma → Group Probs")

            self.input_edit = QLineEdit()
            self.output_edit = QLineEdit()
            self.output_mode_combo = QComboBox()
            self.output_mode_combo.addItems([".csv", ".tsv", "excel"])
            self.lemma_edit = QLineEdit("lemma")
            self.group_edit = QLineEdit()
            self.encoding_edit = QLineEdit("utf-8")
            self.meta_path: Optional[Path] = None
            self.second_spin = QSpinBox()
            self.second_spin.setRange(0, 100)
            self.second_spin.setValue(50)
            self.second_spin.setSuffix(" %")
            self.overlap_combo = QComboBox()
            self.overlap_combo.addItems(["intersection", "correlation", "cosine"])
            self.auto_groups_check = QCheckBox("Auto group")
            self.importance_cutoff_edit = QLineEdit("0.0")
            self.ignore_ambiguous_check = QCheckBox("Ignore ambiguous verbs")
            self.ignore_ambiguous_check.setChecked(True)
            self.importance_cutoff_edit.setEnabled(False)
            self.ignore_ambiguous_check.setEnabled(False)
            self.vba_links_check = QCheckBox("VBA links (.xlsm)")
            self.vba_links_check.setToolTip(
                "Writes .xlsm output and reserves VBA link mode. Current export still uses formula-based hyperlinks."
            )
            self.progress_bar = QProgressBar()
            self.progress_bar.setRange(0, 100)
            self.progress_bar.setValue(0)
            self.progress_bar.setTextVisible(True)
            self.progress_bar.setFormat("Idle")
            self.progress_label = QLabel("Ready")
            self.progress_label.setStyleSheet("color: #666; font-style: italic;")
            self.progress_text = QTextEdit()
            self.progress_text.setReadOnly(True)
            self.progress_text.setMaximumHeight(200)

            browse_in = QPushButton("Browse...")
            browse_out = QPushButton("Browse...")
            load_json_btn = QPushButton("Load Settings from JSON")
            run_btn = QPushButton("Run")

            browse_in.clicked.connect(self.pick_input)
            browse_out.clicked.connect(self.pick_output)
            load_json_btn.clicked.connect(self.pick_metadata)
            run_btn.clicked.connect(self.start)
            self.auto_groups_check.toggled.connect(self.update_auto_group_controls)
            self.output_mode_combo.currentTextChanged.connect(self.update_output_controls)

            # File paths section
            file_grid = QGridLayout()
            file_grid.addWidget(QLabel("Input CSV:"), 0, 0)
            file_grid.addWidget(self.input_edit, 0, 1)
            file_grid.addWidget(browse_in, 0, 2)
            file_grid.addWidget(QLabel("Output File:"), 1, 0)
            file_grid.addWidget(self.output_edit, 1, 1)
            file_grid.addWidget(browse_out, 1, 2)
            file_grid.addWidget(QLabel("Output Type:"), 2, 0)
            file_grid.addWidget(self.output_mode_combo, 2, 1, 1, 2)

            # Settings section
            settings_box = QGroupBox("Settings")
            settings_layout = QVBoxLayout()
            
            # Load JSON button at top of settings
            metadata_layout = QHBoxLayout()
            metadata_layout.addWidget(load_json_btn)
            metadata_layout.addStretch()
            settings_layout.addLayout(metadata_layout)
            
            # Options in horizontal layout
            options_layout = QHBoxLayout()
            
            lemma_col_layout = QVBoxLayout()
            lemma_col_layout.addWidget(QLabel("Lemma Column:"))
            lemma_col_layout.addWidget(self.lemma_edit)
            
            group_cols_layout = QVBoxLayout()
            group_cols_layout.addWidget(QLabel("Group Cols (comma):"))
            group_cols_layout.addWidget(self.group_edit)
            
            second_layout = QVBoxLayout()
            second_layout.addWidget(QLabel("Second Threshold (% of top):"))
            second_layout.addWidget(self.second_spin)

            overlap_layout = QVBoxLayout()
            overlap_layout.addWidget(QLabel("Matrix Measure:"))
            overlap_layout.addWidget(self.overlap_combo)

            auto_group_layout = QVBoxLayout()
            auto_group_layout.addWidget(self.auto_groups_check)
            auto_group_layout.addWidget(QLabel("Importance Cutoff:"))
            auto_group_layout.addWidget(self.importance_cutoff_edit)
            auto_group_layout.addWidget(self.ignore_ambiguous_check)

            link_layout = QVBoxLayout()
            link_layout.addWidget(QLabel("Link Mode:"))
            link_layout.addWidget(self.vba_links_check)
            
            encoding_layout = QVBoxLayout()
            encoding_layout.addWidget(QLabel("Encoding:"))
            encoding_layout.addWidget(self.encoding_edit)
            
            options_layout.addLayout(lemma_col_layout)
            options_layout.addLayout(group_cols_layout)
            options_layout.addLayout(second_layout)
            options_layout.addLayout(overlap_layout)
            options_layout.addLayout(auto_group_layout)
            options_layout.addLayout(link_layout)
            options_layout.addLayout(encoding_layout)
            
            settings_layout.addLayout(options_layout)
            settings_box.setLayout(settings_layout)

            main = QVBoxLayout()
            main.addLayout(file_grid)
            main.addWidget(settings_box)
            main.addWidget(run_btn)
            main.addWidget(QLabel("Progress"))
            main.addWidget(self.progress_bar)
            main.addWidget(self.progress_label)
            main.addWidget(self.progress_text)

            container = QWidget()
            container.setLayout(main)
            self.setCentralWidget(container)

            self.worker: Optional[Worker] = None

        def pick_input(self):
            path, _ = QFileDialog.getOpenFileName(self, "Select input CSV", "", "CSV Files (*.csv)")
            if path:
                self.input_edit.setText(path)

        def pick_output(self):
            path, _ = QFileDialog.getSaveFileName(self, "Select output", "", "Delimited/Excel Files (*.csv *.tsv *.xlsx *.xlsm)")
            if path:
                self.output_edit.setText(path)
                self.output_mode_combo.setCurrentText(output_mode_from_suffix(Path(path).suffix))

        def pick_metadata(self):
            path, _ = QFileDialog.getOpenFileName(self, "Select metadata JSON", "", "JSON Files (*.json)")
            if not path:
                return
            self.meta_path = Path(path)
            try:
                with Path(path).open("r", encoding="utf-8") as f:
                    meta = json.load(f)
                tool = meta.get("tool")
                if tool == "LemmaToGroupProbs":
                    input_file = meta.get("input_file")
                    if input_file:
                        input_path = Path(str(input_file))
                        if input_path.suffix.lower() == ".csv":
                            self.input_edit.setText(input_file)
                        else:
                            QMessageBox.warning(self, "Metadata error", f"Input file in metadata is not a .csv file: {input_file}")
                    output_file = meta.get("output_file")
                    if output_file:
                        self.output_edit.setText(output_file)
                        self.output_mode_combo.setCurrentText(output_mode_from_suffix(Path(str(output_file)).suffix))
                    settings = meta.get("settings", {}) if isinstance(meta.get("settings"), dict) else {}
                    lemma_col = settings.get("lemma_col")
                    if lemma_col:
                        self.lemma_edit.setText(str(lemma_col))
                    group_cols = settings.get("group_cols")
                    if isinstance(group_cols, list):
                        self.group_edit.setText(", ".join(map(str, group_cols)))
                    second = settings.get("second_threshold")
                    if isinstance(second, (int, float)) and 0.0 <= second <= 1.0:
                        self.second_spin.setValue(int(round(second * 100)))
                    self.auto_groups_check.setChecked(bool(settings.get("auto_groups", False)))
                    importance_cutoff = settings.get("importance_cutoff")
                    if isinstance(importance_cutoff, (int, float)):
                        self.importance_cutoff_edit.setText(str(importance_cutoff))
                    self.ignore_ambiguous_check.setChecked(bool(settings.get("ignore_ambiguous_auto_groups", True)))
                    overlap_measure = settings.get("overlap_measure")
                    if overlap_measure in OVERLAP_MEASURES:
                        self.overlap_combo.setCurrentText(str(overlap_measure))
                    self.vba_links_check.setChecked(bool(settings.get("vba_links", False)))
                    encoding = settings.get("encoding")
                    if encoding:
                        self.encoding_edit.setText(str(encoding))
                    QMessageBox.information(self, "Metadata loaded", "✓ Loaded LemmaToGroupProbs metadata")
                elif tool == "MLMGroupAggregator":
                    # Populate input file, lemma column, group columns (everything except output file)
                    output_file = meta.get("output_file")
                    if output_file:
                        output_path = Path(str(output_file))
                        if output_path.suffix.lower() == ".csv":
                            self.input_edit.setText(str(output_file))
                        else:
                            QMessageBox.warning(self, "Metadata error", f"Output file in metadata is not a .csv file: {output_file}")
                    settings = meta.get("settings", {}) if isinstance(meta.get("settings"), dict) else {}
                    lemma_col = settings.get("lemma_col")
                    if lemma_col:
                        self.lemma_edit.setText(str(lemma_col))
                    groups = meta.get("groups")
                    if isinstance(groups, dict):
                        self.group_edit.setText(", ".join(map(str, groups.keys())))
                    # Also populate max_k from top_k setting if available
                    top_k = settings.get("top_k")
                    if top_k and isinstance(top_k, int) and top_k > 0:
                        # Note: max_k is informational for LemmaToGroupProbs and usually inferred
                        pass
                    QMessageBox.information(self, "Metadata loaded", "✓ Loaded MLMGroupAggregator metadata")
                else:
                    QMessageBox.warning(self, "Metadata error", f"Unknown tool in metadata: {tool}")
            except Exception as e:
                QMessageBox.warning(self, "Metadata error", f"Could not read metadata JSON: {str(e)}")

        def start(self):
            input_path = Path(self.input_edit.text().strip())
            output_path = Path(self.output_edit.text().strip())
            if not input_path.exists():
                QMessageBox.warning(self, "Input missing", "Select a valid input CSV")
                return
            if input_path.suffix.lower() != ".csv":
                QMessageBox.warning(self, "Invalid input", "Input file must be a .csv file")
                return
            if not self.output_edit.text().strip():
                QMessageBox.warning(self, "Output missing", "Select an output filename")
                return
            output_mode = self.output_mode_combo.currentText()
            if output_mode == ".csv":
                output_path = output_path.with_suffix(".csv")
            elif output_mode == ".tsv":
                output_path = output_path.with_suffix(".tsv")
            else:
                output_path = output_path.with_suffix(".xlsm" if self.vba_links_check.isChecked() else ".xlsx")
            self.output_edit.setText(str(output_path))

            lemma_col = self.lemma_edit.text().strip() or "lemma"
            group_cols = [c.strip() for c in self.group_edit.text().split(",") if c.strip()] or None
            second_threshold = self.second_spin.value() / 100.0
            overlap_measure = self.overlap_combo.currentText()
            auto_groups = self.auto_groups_check.isChecked()
            try:
                importance_cutoff = float(self.importance_cutoff_edit.text().strip() or "0")
            except ValueError:
                QMessageBox.warning(self, "Bad cutoff", "Importance cutoff must be a number")
                return
            if importance_cutoff < 0.0:
                QMessageBox.warning(self, "Bad cutoff", "Importance cutoff must be non-negative")
                return
            ignore_ambiguous_auto_groups = self.ignore_ambiguous_check.isChecked()
            use_vba_links = self.vba_links_check.isChecked()
            encoding = self.encoding_edit.text().strip() or "utf-8"

            self.progress_text.clear()
            self.progress_text.append("Starting worker thread...")
            # Determinate placeholder until total is known
            self.progress_bar.setRange(0, 1)
            self.progress_bar.setValue(0)
            self.progress_bar.setFormat("Starting...")
            self.worker = Worker(
                input_path=input_path,
                output_path=output_path,
                lemma_col=lemma_col,
                group_cols=group_cols,
                second_threshold=second_threshold,
                overlap_measure=overlap_measure,
                auto_groups=auto_groups,
                importance_cutoff=importance_cutoff,
                ignore_ambiguous_auto_groups=ignore_ambiguous_auto_groups,
                use_vba_links=use_vba_links,
                encoding=encoding,
                metadata_path=self.meta_path,
            )
            self.worker.progress.connect(self.append_log)
            self.worker.progress_value.connect(self.update_progress)
            self.worker.finished.connect(self.done)
            self.worker.start()

        def append_log(self, msg: str):
            self.progress_text.append(msg)

        def update_auto_group_controls(self, checked: bool):
            excel_enabled = self.output_mode_combo.currentText() == "excel"
            self.importance_cutoff_edit.setEnabled(checked and excel_enabled)
            self.ignore_ambiguous_check.setEnabled(checked and excel_enabled)

        def update_output_controls(self, _mode: str):
            excel_enabled = self.output_mode_combo.currentText() == "excel"
            self.overlap_combo.setEnabled(excel_enabled)
            self.vba_links_check.setEnabled(excel_enabled)
            self.auto_groups_check.setEnabled(excel_enabled)
            self.update_auto_group_controls(self.auto_groups_check.isChecked())

        def update_progress(self, current: int, total: int):
            if total and total > 0:
                if self.progress_bar.maximum() != total:
                    self.progress_bar.setRange(0, total)
                self.progress_bar.setValue(current)
                self.progress_bar.setFormat(f"{current}/{total}")
            else:
                if self.progress_bar.maximum() != 1:
                    self.progress_bar.setRange(0, 1)
                self.progress_bar.setValue(0)
                self.progress_bar.setFormat("Working...")

        def done(self, ok: bool, msg: str):
            self.progress_bar.setRange(0, 100)
            self.progress_bar.setValue(100)
            self.progress_bar.setFormat("Done" if ok else "Error")
            self.progress_label.setText("Complete" if ok else "Failed")
            if ok:
                QMessageBox.information(self, "Done", msg)
            else:
                QMessageBox.critical(self, "Error", msg)

    app = QApplication(sys.argv)
    window = MainWindow()
    window.resize(640, 480)
    window.show()
    app.exec()


def main():
    if len(sys.argv) == 1:
        run_gui()
    else:
        run_cli()


if __name__ == "__main__":
    main()
