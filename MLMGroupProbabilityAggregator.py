#!/usr/bin/env python3
"""
Combined MLM Group Aggregation → Lemma-to-Group Probability conversion.

Supports both CLI and GUI modes:
  - Run without arguments: launches GUI (requires PySide6)
  - Run with arguments: uses CLI mode

Pipeline:
  1. Aggregate MLM predictions by group (like MLMGroupAggregator)
  2. Convert group probabilities to lemma-based summary (like LemmaToGroupProbs)
  3. Output Excel file with two sheets:
     - lemma_to_groups: lemmas × group probabilities (formatted as %)
     - groups_ranked: each group's top lemmas sorted by probability
  4. Optionally save intermediate group aggregation CSV

Output format:
  - Default: Excel file (.xlsx) with two sheets
  - With --keep-csv: also save intermediate aggregated CSV
  - With --csv-only: save only the intermediate CSV (no Excel)

Supports metadata loading from:
  - RoBERTaMaskedLanguageModelVerbs.json (auto-uses output as MLM input)
  - MLMGroupAggregator.json (loads all settings)

Requirements:
  pip install lemminflect openpyxl
  pip install PySide6  # for GUI mode only

CLI Examples:
  python MLMGroupProbabilityAggregator.py mlm_out.csv groups.csv output.xlsx
  python MLMGroupProbabilityAggregator.py mlm_out.csv groups.csv output.xlsx --keep-csv
  python MLMGroupProbabilityAggregator.py mlm_out.csv groups.csv output.xlsx --csv-only
  python MLMGroupProbabilityAggregator.py --load-metadata mlm_out.json mlm_out.csv groups.csv output.xlsx

GUI Mode:
  python MLMGroupProbabilityAggregator.py
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import logging
import re
import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple, Sequence

from lemminflect import getLemma

PROB_COL_RE = re.compile(r"^prob_(\d+)$")


# ============================================================================
# VERSION TRACKING AND METADATA
# ============================================================================

def get_aggregator_version_info() -> Dict:
    """Get version information for dependencies."""
    try:
        import lemminflect
        lemminflect_version = lemminflect.__version__
    except (ImportError, AttributeError):
        lemminflect_version = "unknown"
    
    try:
        import openpyxl
        openpyxl_version = openpyxl.__version__
    except (ImportError, AttributeError):
        openpyxl_version = "unknown"
    
    return {
        "lemminflect": lemminflect_version,
        "openpyxl": openpyxl_version,
    }


def compute_file_md5(path: Path) -> str:
    """Compute MD5 checksum of a file."""
    md5 = hashlib.md5()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            md5.update(chunk)
    return md5.hexdigest()


def save_combined_metadata(
    output_path: Path,
    mlm_csv_path: Path,
    group_csv_path: Path,
    mlm_checksum: str,
    group_checksum: str,
    output_checksum: str,
    settings: Dict[str, Any],
    stats: Dict[str, Any],
    group_labels: List[str],
    lemma_to_groups: Dict[str, Set[str]],
    source_metadata: Optional[Dict] = None,
) -> None:
    """Save combined aggregation metadata to JSON file."""
    json_path = output_path.with_suffix(".json")
    
    groups_structure = {}
    for group_name in group_labels:
        lemmas_in_group = sorted([
            lemma for lemma, groups in lemma_to_groups.items()
            if group_name in groups
        ])
        groups_structure[group_name] = lemmas_in_group
    
    metadata = {
        "timestamp": datetime.utcnow().isoformat() + "Z",
        "tool": "MLMGroupProbabilityAggregator",
        "versions": get_aggregator_version_info(),
        "output_file": str(output_path),
        "output_checksum": output_checksum,
        "input_files": {
            "mlm_csv": str(mlm_csv_path),
            "group_csv": str(group_csv_path),
        },
        "input_checksums": {
            "mlm_csv": mlm_checksum,
            "group_csv": group_checksum,
        },
        "groups": groups_structure,
        "settings": {
            "top_k": settings.get("top_k"),
            "lemma_col": settings.get("lemma_col", "lemma"),
            "short": settings.get("short", False),
            "include_count": settings.get("include_count", False),
            "encoding": settings.get("encoding", "utf-8-sig"),
            "second_threshold": settings.get("second_threshold", 0.50),
            "keep_csv": settings.get("keep_csv", False),
            "csv_only": settings.get("csv_only", False),
        },
        "statistics": stats,
    }
    
    if source_metadata:
        metadata["source_metadata"] = source_metadata
    
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(metadata, f, indent=2)


def load_combined_metadata(json_path: Path) -> Any:
    """Load metadata from JSON file."""
    with json_path.open("r", encoding="utf-8") as f:
        return json.load(f)


# ============================================================================
# AGGREGATION LOGIC (from MLMGroupAggregator)
# ============================================================================

def load_groups_csv(group_csv_path: str, encoding: str, logger: logging.Logger) -> Tuple[List[str], Dict[str, Set[str]]]:
    """Load group definitions from CSV: columns=group names, cells=lemmas."""
    group_labels: List[str] = []
    lemma_to_groups: Dict[str, Set[str]] = defaultdict(set)
    
    try:
        with open(group_csv_path, "r", encoding=encoding) as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                raise ValueError("Group CSV has no header")
            
            group_labels = list(reader.fieldnames)
            
            for row_num, row in enumerate(reader, start=2):
                for group_name in group_labels:
                    cell = row.get(group_name, "").strip()
                    if cell:
                        for lemma in cell.split(","):
                            lemma = lemma.strip()
                            if lemma:
                                lemma_to_groups[lemma].add(group_name)
    except Exception as e:
        logger.error(f"Error loading group CSV: {e}")
        raise
    
    return group_labels, lemma_to_groups


def infer_top_k(fieldnames: Sequence[str]) -> int:
    """Infer top_k from fieldnames by counting prob_i columns."""
    max_k = 0
    for name in fieldnames:
        match = PROB_COL_RE.match(name)
        if match:
            k = int(match.group(1))
            max_k = max(max_k, k)
    return max_k


def normalize_pred_token(tok: str) -> str:
    """Normalize a prediction token: remove special chars, lowercase."""
    return re.sub(r"[^\w]", "", tok).lower()


def aggregate_and_output(
    mlm_csv_path: Path,
    output_csv_path: Optional[Path],
    group_labels: List[str],
    lemma_to_groups: Dict[str, Set[str]],
    settings: Dict[str, Any],
    logger: logging.Logger,
) -> Tuple[Dict[str, Dict[str, float]], Dict[str, int], int, int]:
    """
    Aggregate MLM probabilities by group and write intermediate CSV if requested.
    
    Returns: (means, counts, processed, skipped)
    """
    lemma_col = settings.get("lemma_col", "lemma")
    top_k = settings.get("top_k", 0)
    include_count = settings.get("include_count", False)
    encoding = settings.get("encoding", "utf-8-sig")
    
    means: Dict[str, Dict[str, float]] = {g: defaultdict(float) for g in group_labels}
    counts: Dict[str, int] = defaultdict(int)
    processed = 0
    skipped = 0
    
    reader_fieldnames: Optional[List[str]] = None
    output_fieldnames: Optional[List[str]] = None
    output_writer = None
    output_file = None
    
    try:
        with open(mlm_csv_path, "r", encoding=encoding) as infile:
            reader = csv.DictReader(infile)
            reader_fieldnames = reader.fieldnames or []
            
            # Open output file if requested
            if output_csv_path:
                output_file = open(output_csv_path, "w", encoding=encoding, newline="")
                output_fieldnames = list(reader_fieldnames) + group_labels
                if include_count:
                    output_fieldnames.append("group_count")
                output_writer = csv.DictWriter(output_file, fieldnames=output_fieldnames)
                output_writer.writeheader()
            
            for row in reader:
                lemma = row.get(lemma_col, "").strip()
                if not lemma:
                    skipped += 1
                    continue
                
                # Normalize lemma
                lemma_norm = normalize_pred_token(lemma)
                
                # Lemmatize if not already a lemma
                try:
                    lemmas = getLemma(lemma_norm, pos="VERB")
                    lemma = lemmas[0] if lemmas else lemma_norm
                except:
                    lemma = lemma_norm
                
                # Aggregate probabilities by group
                group_probs: Dict[str, float] = {g: 0.0 for g in group_labels}
                for i in range(1, top_k + 1):
                    token_col = f"token_{i}"
                    prob_col = f"prob_{i}"
                    
                    token = row.get(token_col, "").strip()
                    try:
                        prob = float(row.get(prob_col, 0))
                    except ValueError:
                        prob = 0
                    
                    if token and prob > 0:
                        token_norm = normalize_pred_token(token)
                        for group_name in group_labels:
                            if token_norm in lemma_to_groups and group_name in lemma_to_groups[token_norm]:
                                group_probs[group_name] += prob
                
                # Update means
                counts[lemma] += 1
                for group_name in group_labels:
                    means[group_name][lemma] += group_probs[group_name]
                
                # Write output row if needed
                if output_writer:
                    out_row = dict(row)
                    for group_name in group_labels:
                        out_row[group_name] = group_probs[group_name]
                    if include_count:
                        out_row["group_count"] = len([g for g in group_probs if group_probs[g] > 0])
                    output_writer.writerow(out_row)
                
                processed += 1
                if processed % 10000 == 0:
                    logger.info(f"Processed {processed:,} rows...")
    
    finally:
        if output_file:
            output_file.close()
    
    # Compute means
    for group_name in group_labels:
        for lemma in means[group_name]:
            if counts[lemma] > 0:
                means[group_name][lemma] /= counts[lemma]
    
    return means, counts, processed, skipped


# ============================================================================
# EXCEL OUTPUT LOGIC (from LemmaToGroupProbs)
# ============================================================================

def write_excel_output(
    output_xlsx_path: Path,
    means: Dict[str, Dict[str, float]],
    counts: Dict[str, int],
    group_labels: List[str],
    second_threshold: float,
    logger: logging.Logger,
) -> None:
    """Write Excel output with two sheets: lemma_to_groups and groups_ranked."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit("openpyxl required for Excel output. Install with: pip install openpyxl")
    
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "lemma_to_groups"
    
    # Headers
    ws1.append(["lemma"] + group_labels)
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws1[1]:
        cell.font = header_font
        cell.alignment = header_align
    
    bold_font = Font(bold=True)
    blue_font = Font(bold=True, color="0070C0")
    
    # Data rows
    best_group_for_lemma: Dict[str, str] = {}
    for lemma in sorted(counts.keys()):
        row = [lemma]
        for g in group_labels:
            row.append(means[g].get(lemma, 0.0))
        ws1.append(row)
        
        r = ws1.max_row
        lemma_cell = ws1.cell(row=r, column=1)
        
        # Find best and 2nd-best groups
        group_vals = [(g, means[g].get(lemma, 0.0)) for g in group_labels]
        group_vals.sort(key=lambda x: x[1], reverse=True)
        
        if group_vals:
            max_g, max_v = group_vals[0]
            second_v = group_vals[1][1] if len(group_vals) > 1 else 0
            best_group_for_lemma[lemma] = max_g
            
            # Bold the max group cell
            max_idx = group_labels.index(max_g)
            max_col = 2 + max_idx
            ws1.cell(row=r, column=max_col).font = bold_font
            
            # Color lemma blue if 2nd-best >= threshold * best
            if max_v > 0 and second_v >= (second_threshold * max_v):
                lemma_cell.font = blue_font
    
    # Format group columns as percentages
    for j in range(2, 2 + len(group_labels)):
        for i in range(2, ws1.max_row + 1):
            ws1.cell(row=i, column=j).number_format = "0.00%"
    
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(ws1.max_column)}{ws1.max_row}"
    
    # Sheet 2: groups_ranked
    ws2 = wb.create_sheet(title="groups_ranked")
    
    per_group_sorted: Dict[str, List[Tuple[str, float]]] = {}
    max_len = 0
    for g in group_labels:
        items = [(lemma, means[g].get(lemma, 0.0)) for lemma in counts]
        items.sort(key=lambda x: x[1], reverse=True)
        per_group_sorted[g] = items
        max_len = max(max_len, len(items))
    
    header2 = []
    for g in group_labels:
        header2.extend([f"{g}_lemma", f"{g}_pct"])
    ws2.append(header2)
    for cell in ws2[1]:
        cell.font = header_font
        cell.alignment = header_align
    
    for r in range(max_len):
        ws2.append([None] * (2 * len(group_labels)))
        excel_row = ws2.max_row
        
        for k, g in enumerate(group_labels):
            if r < len(per_group_sorted[g]):
                lemma, pct = per_group_sorted[g][r]
                lemma_col_idx = 1 + 2 * k
                pct_col_idx = 2 + 2 * k
                
                cell_lemma = ws2.cell(row=excel_row, column=lemma_col_idx, value=lemma)
                ws2.cell(row=excel_row, column=pct_col_idx, value=pct)
                
                if best_group_for_lemma.get(lemma) == g:
                    cell_lemma.font = bold_font
    
    for col in range(2, 2 * len(group_labels) + 1, 2):
        for i in range(2, ws2.max_row + 1):
            ws2.cell(row=i, column=col).number_format = "0.00%"
    
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(ws2.max_column)}{ws2.max_row}"
    
    wb.save(output_xlsx_path)
    logger.info(f"Wrote Excel output: {output_xlsx_path}")


# ============================================================================
# CLI MODE
# ============================================================================

def run_cli() -> None:
    """Run in CLI mode."""
    ap = argparse.ArgumentParser(
        description="Combine MLM group aggregation with lemma-to-group probability conversion"
    )
    ap.add_argument("mlm_csv", help="MLM output CSV (from RoBERTaMaskedLanguageModelVerbs)")
    ap.add_argument("group_csv", help="Group definitions CSV")
    ap.add_argument("output", help="Output file (.xlsx or .csv)")
    ap.add_argument("--load-metadata", help="Load settings from metadata JSON")
    ap.add_argument("--top-k", type=int, default=0, help="Top-k predictions (default: infer from CSV)")
    ap.add_argument("--lemma-col", default="lemma", help="Lemma column name (default: lemma)")
    ap.add_argument("--second-threshold", type=float, default=0.50, help="Ambiguity threshold (default: 0.50)")
    ap.add_argument("--include-count", action="store_true", help="Include group count in output")
    ap.add_argument("--keep-csv", action="store_true", help="Keep intermediate CSV alongside output")
    ap.add_argument("--csv-only", action="store_true", help="Output only CSV (no Excel)")
    ap.add_argument("--encoding", default="utf-8-sig", help="File encoding (default: utf-8-sig)")
    ap.add_argument("--log-level", default="INFO", help="Logging level")
    args = ap.parse_args()
    
    logging.basicConfig(
        level=getattr(logging, args.log_level.upper(), logging.INFO),
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )
    logger = logging.getLogger("agg_prob")
    
    mlm_path = Path(args.mlm_csv)
    group_path = Path(args.group_csv)
    output_path = Path(args.output)
    
    if not mlm_path.exists():
        raise SystemExit(f"MLM CSV not found: {mlm_path}")
    if not group_path.exists():
        raise SystemExit(f"Group CSV not found: {group_path}")
    
    # Load metadata if provided
    source_metadata = None
    if args.load_metadata:
        metadata_path = Path(args.load_metadata)
        try:
            metadata = load_combined_metadata(metadata_path)
            tool = metadata.get("tool", "unknown")
            
            if tool == "RoBERTaMaskedLanguageModelVerbs":
                mlm_path = Path(metadata.get("output_file", mlm_path))
            elif tool in ["MLMGroupAggregator", "MLMGroupProbabilityAggregator"]:
                input_files = metadata.get("input_files", {})
                mlm_path = Path(input_files.get("mlm_csv", mlm_path))
                group_path = Path(input_files.get("group_csv", group_path))
            
            settings = metadata.get("settings", {})
            args.top_k = args.top_k or settings.get("top_k", 0)
            args.lemma_col = args.lemma_col or settings.get("lemma_col", "lemma")
            args.second_threshold = settings.get("second_threshold", args.second_threshold)
            args.include_count = settings.get("include_count", args.include_count)
            args.keep_csv = settings.get("keep_csv", args.keep_csv)
            args.csv_only = settings.get("csv_only", args.csv_only)
            args.encoding = settings.get("encoding", args.encoding)
            
            source_metadata = metadata.get("source_metadata")
            logger.info(f"Loaded settings from {metadata_path}")
        except Exception as e:
            logger.error(f"Failed to load metadata: {e}")
            raise SystemExit(str(e))
    
    logger.info(f"MLM CSV: {mlm_path}")
    logger.info(f"Group CSV: {group_path}")
    logger.info(f"Output: {output_path}")
    
    # Load groups
    group_labels, lemma_to_groups = load_groups_csv(str(group_path), args.encoding, logger)
    logger.info(f"Loaded {len(group_labels)} groups with {len(lemma_to_groups)} lemmas")
    
    # Infer top_k if not provided
    if args.top_k <= 0:
        with open(mlm_path, "r", encoding=args.encoding) as f:
            reader = csv.DictReader(f)
            if reader.fieldnames:
                args.top_k = infer_top_k(reader.fieldnames)
        logger.info(f"Inferred top_k={args.top_k}")
    
    settings = {
        "top_k": args.top_k,
        "lemma_col": args.lemma_col,
        "short": False,
        "include_count": args.include_count,
        "encoding": args.encoding,
        "second_threshold": args.second_threshold,
        "keep_csv": args.keep_csv,
        "csv_only": args.csv_only,
    }
    
    # Compute intermediate CSV path
    intermediate_csv = None
    if args.keep_csv or args.csv_only:
        if args.csv_only and output_path.suffix.lower() == ".xlsx":
            intermediate_csv = output_path.with_stem(output_path.stem + "_intermediate")
        else:
            intermediate_csv = output_path if args.csv_only else output_path.with_suffix(".csv")
    
    # Run aggregation
    start_time = time.time()
    means, counts, processed, skipped = aggregate_and_output(
        mlm_path, intermediate_csv, group_labels, lemma_to_groups, settings, logger
    )
    
    # Write Excel if requested
    if not args.csv_only:
        write_excel_output(output_path, means, counts, group_labels, args.second_threshold, logger)
    
    elapsed = time.time() - start_time
    logger.info(f"Processed {processed:,} rows ({skipped} skipped) in {elapsed:.1f}s")
    
    # Compute checksums and save metadata
    mlm_checksum = compute_file_md5(mlm_path)
    group_checksum = compute_file_md5(group_path)
    output_checksum = compute_file_md5(output_path if not args.csv_only else intermediate_csv)
    
    stats = {
        "processed_rows": processed,
        "skipped_rows": skipped,
        "unique_lemmas": len(counts),
        "elapsed_seconds": elapsed,
    }
    
    if not args.csv_only:
        save_combined_metadata(
            output_path, mlm_path, group_path, mlm_checksum, group_checksum,
            output_checksum, settings, stats, group_labels, lemma_to_groups, source_metadata
        )
        logger.info(f"Wrote metadata: {output_path.with_suffix('.json')}")


# ============================================================================
# GUI MODE
# ============================================================================

def run_gui() -> None:
    """Launch the GUI application."""
    try:
        from PySide6.QtCore import Qt, QThread, Signal
        from PySide6.QtWidgets import (
            QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
            QGroupBox, QLabel, QLineEdit, QPushButton, QFileDialog, QTextEdit,
            QProgressBar, QCheckBox, QDoubleSpinBox, QMessageBox, QSpinBox,
        )
    except ImportError:
        print("Error: PySide6 is required for GUI mode.")
        print("Install it with: pip install PySide6")
        sys.exit(1)
    
    class AggregationWorker(QThread):
        """Worker thread for aggregation without blocking UI."""
        progress = Signal(str)
        finished = Signal(bool, str)
        
        def __init__(self, mlm_path: Path, group_path: Path, output_path: Path, settings: Dict[str, Any]):
            super().__init__()
            self.mlm_path = mlm_path
            self.group_path = group_path
            self.output_path = output_path
            self.settings = settings
        
        def run(self):
            try:
                logger = logging.getLogger("worker")
                
                # Load groups
                group_labels, lemma_to_groups = load_groups_csv(
                    str(self.group_path), self.settings["encoding"], logger
                )
                self.progress.emit(f"Loaded {len(group_labels)} groups")
                
                # Infer top_k if needed
                if self.settings["top_k"] <= 0:
                    with open(self.mlm_path, "r", encoding=self.settings["encoding"]) as f:
                        reader = csv.DictReader(f)
                        if reader.fieldnames:
                            self.settings["top_k"] = infer_top_k(reader.fieldnames)
                
                # Aggregate
                intermediate_csv = None
                if self.settings["keep_csv"] or self.settings["csv_only"]:
                    if self.settings["csv_only"] and self.output_path.suffix.lower() == ".xlsx":
                        intermediate_csv = self.output_path.with_stem(self.output_path.stem + "_intermediate")
                    else:
                        intermediate_csv = self.output_path if self.settings["csv_only"] else self.output_path.with_suffix(".csv")
                
                means, counts, processed, skipped = aggregate_and_output(
                    self.mlm_path, intermediate_csv, group_labels, lemma_to_groups, self.settings, logger
                )
                
                self.progress.emit(f"Aggregated {processed:,} rows")
                
                # Write Excel
                if not self.settings["csv_only"]:
                    write_excel_output(
                        self.output_path, means, counts, group_labels,
                        self.settings["second_threshold"], logger
                    )
                    self.progress.emit("Excel output written")
                
                self.finished.emit(True, "Complete!")
            except Exception as e:
                self.progress.emit(f"Error: {e}")
                self.finished.emit(False, str(e))
    
    class MainWindow(QMainWindow):
        """Main GUI window."""
        
        def __init__(self):
            super().__init__()
            self.setWindowTitle("MLM Group Probability Aggregator")
            self.worker = None
            self.init_ui()
        
        def init_ui(self):
            """Initialize UI components."""
            widget = QWidget()
            layout = QVBoxLayout()
            
            # Input files
            input_group = QGroupBox("Input Files")
            input_layout = QVBoxLayout()
            
            input_layout.addWidget(QLabel("MLM CSV:"))
            self.mlm_input = QLineEdit()
            mlm_btn = QPushButton("Browse...")
            mlm_btn.clicked.connect(lambda: self.browse_file(self.mlm_input, "Select MLM CSV"))
            input_layout.addLayout(self._h_layout(self.mlm_input, mlm_btn))
            
            input_layout.addWidget(QLabel("Group CSV:"))
            self.group_input = QLineEdit()
            group_btn = QPushButton("Browse...")
            group_btn.clicked.connect(lambda: self.browse_file(self.group_input, "Select Group CSV"))
            input_layout.addLayout(self._h_layout(self.group_input, group_btn))
            
            input_group.setLayout(input_layout)
            layout.addWidget(input_group)
            
            # Output
            output_group = QGroupBox("Output")
            output_layout = QVBoxLayout()
            
            output_layout.addWidget(QLabel("Output File (.xlsx or .csv):"))
            self.output_input = QLineEdit()
            output_btn = QPushButton("Browse...")
            output_btn.clicked.connect(lambda: self.browse_file(self.output_input, "Select Output"))
            output_layout.addLayout(self._h_layout(self.output_input, output_btn))
            
            output_group.setLayout(output_layout)
            layout.addWidget(output_group)
            
            # Settings
            settings_group = QGroupBox("Settings")
            settings_layout = QVBoxLayout()
            
            settings_layout.addWidget(QLabel("Lemma Column:"))
            self.lemma_col = QLineEdit("lemma")
            settings_layout.addWidget(self.lemma_col)
            
            settings_layout.addWidget(QLabel("Top-K:"))
            self.topk_spin = QSpinBox()
            self.topk_spin.setValue(0)
            self.topk_spin.setMinimum(0)
            settings_layout.addWidget(self.topk_spin)
            
            settings_layout.addWidget(QLabel("Ambiguity Threshold:"))
            self.threshold_spin = QDoubleSpinBox()
            self.threshold_spin.setValue(0.50)
            self.threshold_spin.setMinimum(0.0)
            self.threshold_spin.setMaximum(1.0)
            self.threshold_spin.setSingleStep(0.05)
            settings_layout.addWidget(self.threshold_spin)
            
            self.count_check = QCheckBox("Include group count")
            settings_layout.addWidget(self.count_check)
            
            self.keep_csv_check = QCheckBox("Keep intermediate CSV")
            settings_layout.addWidget(self.keep_csv_check)
            
            self.csv_only_check = QCheckBox("CSV only (no Excel)")
            settings_layout.addWidget(self.csv_only_check)
            
            settings_group.setLayout(settings_layout)
            layout.addWidget(settings_group)
            
            # Progress
            self.progress_bar = QProgressBar()
            self.progress_bar.setMaximum(0)
            layout.addWidget(self.progress_bar)
            
            # Log
            self.log_text = QTextEdit()
            self.log_text.setReadOnly(True)
            layout.addWidget(self.log_text)
            
            # Buttons
            button_layout = QHBoxLayout()
            self.start_btn = QPushButton("Start")
            self.start_btn.clicked.connect(self.start_aggregation)
            load_btn = QPushButton("Load Metadata")
            load_btn.clicked.connect(self.load_metadata)
            button_layout.addWidget(load_btn)
            button_layout.addWidget(self.start_btn)
            layout.addLayout(button_layout)
            
            widget.setLayout(layout)
            self.setCentralWidget(widget)
            self.resize(700, 900)
        
        def _h_layout(self, widget, button):
            layout = QHBoxLayout()
            layout.addWidget(widget)
            layout.addWidget(button)
            return layout
        
        def browse_file(self, line_edit, title):
            path, _ = QFileDialog.getOpenFileName(self, title)
            if path:
                line_edit.setText(path)
        
        def load_metadata(self):
            path, _ = QFileDialog.getOpenFileName(self, "Select Metadata JSON", "", "JSON Files (*.json)")
            if path:
                try:
                    metadata = load_combined_metadata(Path(path))
                    tool = metadata.get("tool")
                    input_files = metadata.get("input_files", {})
                    settings = metadata.get("settings", {})
                    
                    if tool == "RoBERTaMaskedLanguageModelVerbs":
                        self.mlm_input.setText(metadata.get("output_file", ""))
                    elif tool in ["MLMGroupAggregator", "MLMGroupProbabilityAggregator"]:
                        self.mlm_input.setText(input_files.get("mlm_csv", ""))
                        self.group_input.setText(input_files.get("group_csv", ""))
                        self.output_input.setText(metadata.get("output_file", ""))
                    
                    self.lemma_col.setText(settings.get("lemma_col", "lemma"))
                    self.topk_spin.setValue(settings.get("top_k", 0))
                    self.threshold_spin.setValue(settings.get("second_threshold", 0.50))
                    self.count_check.setChecked(settings.get("include_count", False))
                    self.keep_csv_check.setChecked(settings.get("keep_csv", False))
                    self.csv_only_check.setChecked(settings.get("csv_only", False))
                    
                    self.log(f"✓ Loaded metadata from {path}")
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Failed to load metadata: {e}")
        
        def start_aggregation(self):
            mlm = Path(self.mlm_input.text().strip())
            group = Path(self.group_input.text().strip())
            output = Path(self.output_input.text().strip())
            
            if not mlm.exists() or not group.exists() or not self.output_input.text():
                QMessageBox.warning(self, "Input Error", "Fill in all required fields")
                return
            
            settings = {
                "top_k": self.topk_spin.value(),
                "lemma_col": self.lemma_col.text() or "lemma",
                "second_threshold": self.threshold_spin.value(),
                "include_count": self.count_check.isChecked(),
                "keep_csv": self.keep_csv_check.isChecked(),
                "csv_only": self.csv_only_check.isChecked(),
                "encoding": "utf-8-sig",
            }
            
            self.log_text.clear()
            self.start_btn.setEnabled(False)
            self.worker = AggregationWorker(mlm, group, output, settings)
            self.worker.progress.connect(self.log)
            self.worker.finished.connect(self.on_finished)
            self.worker.start()
        
        def log(self, msg: str):
            self.log_text.append(msg)
            self.log_text.verticalScrollBar().setValue(
                self.log_text.verticalScrollBar().maximum()
            )
        
        def on_finished(self, success: bool, msg: str):
            self.start_btn.setEnabled(True)
            if success:
                QMessageBox.information(self, "Success", msg)
            else:
                QMessageBox.critical(self, "Error", msg)
    
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

def main():
    """Main entry point - choose GUI or CLI mode."""
    if len(sys.argv) == 1:
        run_gui()
    else:
        run_cli()


if __name__ == "__main__":
    main()
